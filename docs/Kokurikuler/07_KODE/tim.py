import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import RadarChart, Reference
from openpyxl.utils import get_column_letter
OUT="/sessions/vibrant-inspiring-einstein/mnt/Ko Kurikuler/SAKOLA_WALUYA_XI_2026/02_UNTUK_MURID"
NAVY='1F3864'; GRN='2E7D32'; GOLD='BF8F00'; ROSE='AD1457'; ORG='C55A11'; BLUE='2E74B5'
thin=Side(style='thin',color='B0BEC5'); BD=Border(left=thin,right=thin,top=thin,bottom=thin)
FILL={'GRN':'E8F5E9','GOLD':'FFF6E0','ROSE':'FDE8EF','NAVY':'E8F0FA','ORG':'FCE4D6'}
IN=PatternFill('solid',fgColor='FFF9C4')  # kuning = diisi murid

def head(ws,t,s,col,n=8):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n)
    c=ws.cell(1,1,t); c.font=Font(bold=True,size=16,color='FFFFFF'); c.fill=PatternFill('solid',fgColor=col)
    c.alignment=Alignment('left','center'); ws.row_dimensions[1].height=30
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=n)
    c=ws.cell(2,1,s); c.font=Font(italic=True,size=11,color='595959'); c.alignment=Alignment(wrap_text=True,vertical='top')
    ws.row_dimensions[2].height=32
def th(ws,r,cols,col):
    for i,x in enumerate(cols,1):
        c=ws.cell(r,i,x); c.font=Font(bold=True,color='FFFFFF',size=11); c.fill=PatternFill('solid',fgColor=col)
        c.alignment=Alignment('center','center',wrap_text=True); c.border=BD
    ws.row_dimensions[r].height=32
def w(ws,ws_):
    for i,x in enumerate(ws_,1): ws.column_dimensions[get_column_letter(i)].width=x
def blank(ws,r0,r1,c0,c1):
    for r in range(r0,r1+1):
        for c in range(c0,c1+1):
            x=ws.cell(r,c); x.fill=IN; x.border=BD; x.alignment=Alignment(vertical='center',wrap_text=True); x.font=Font(size=11)
        ws.row_dimensions[r].height=26

wb=openpyxl.Workbook()

# ===== 0. IDENTITAS =====
ws=wb.active; ws.title='0. IDENTITAS TIM'
head(ws,'TEMPLATE DIGITAL TIM — SAKOLA WALUYA','Satu file untuk SATU TIM. Semua kotak KUNING diisi kalian. Yang ditulis di BUKU TULIS hanya: JURNAL REFLEKSI harian + KONTRAK KEBIASAAN.',NAVY,6)
w(ws,[26,30,26,26,20,20])
rows=[('KELAS',''),('NAMA TIM',''),('ZONA AKSI',''),('SEKTOR KAMI',''),('GELOMBANG (BIRU / JINGGA)',''),('WALI KELAS','')]
r=4
for a,b in rows:
    c=ws.cell(r,1,a); c.font=Font(bold=True,size=11); c.fill=PatternFill('solid',fgColor=FILL['NAVY']); c.border=BD
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
    ws.cell(r,2).fill=IN
    for i in range(2,5): ws.cell(r,i).border=BD
    ws.row_dimensions[r].height=26; r+=1
r+=1
c=ws.cell(r,1,'ANGGOTA TIM & PERANNYA'); c.font=Font(bold=True,size=12,color=NAVY); r+=1
th(ws,r,['No','NAMA','PERAN DI KELAS','Apa tugasku dengan peran itu? (tulis sendiri)','',''],BLUE); r+=1
for i in range(6):
    ws.cell(r,1,i+1).border=BD; ws.cell(r,1).alignment=Alignment('center')
    blank(ws,r,r,2,4); r+=1
r+=1
c=ws.cell(r,1,'PERAN: KAPTEN TIM · TIM WAKTU · TIM PIKET DATA · TIM K3 · TIM DOKUMENTASI · TIM LOGISTIK')
c.font=Font(size=11,italic=True,color=GREY) if False else Font(size=11,italic=True,color='595959')

# ===== 1. HARI 1 RADAR =====
ws=wb.create_sheet('H1 · RADAR KEBIASAAN')
head(ws,'HARI 1 · CAGEUR — RADAR KEBIASAAN','Setiap anggota tim mengisi skornya sendiri (1–10). Grafik radar akan muncul otomatis. Jujur — tidak ada yang menghukummu.',GRN,9)
w(ws,[24,15,12,12,12,12,12,12,16])
th(ws,4,['KEBIASAAN (7 KAIH)','SKOR SEKOLAH','Anggota 1','Anggota 2','Anggota 3','Anggota 4','Anggota 5','Anggota 6','RATA-RATA TIM'],GRN)
KEB=[('Bangun Pagi',7.93),('Beribadah',7.34),('Berolahraga',5.92),('Makan Sehat & Bergizi',6.44),('Gemar Belajar',5.50),('Bermasyarakat',6.41),('Tidur Cepat',6.41)]
r=5
for k,v in KEB:
    ws.cell(r,1,k).font=Font(bold=True,size=11); ws.cell(r,1).border=BD; ws.cell(r,1).fill=PatternFill('solid',fgColor=FILL['GRN'])
    c=ws.cell(r,2,v); c.border=BD; c.alignment=Alignment('center'); c.font=Font(size=11,color='C00000',bold=True)
    blank(ws,r,r,3,8)
    c=ws.cell(r,9,f'=IFERROR(ROUND(AVERAGE(C{r}:H{r}),2),"")'); c.border=BD; c.alignment=Alignment('center'); c.font=Font(bold=True,size=11)
    r+=1
ch=RadarChart(); ch.type='marker'; ch.style=26; ch.title='RADAR KEBIASAAN TIM KAMI'
data=Reference(ws,min_col=2,max_col=2,min_row=4,max_row=11)
data2=Reference(ws,min_col=9,max_col=9,min_row=4,max_row=11)
cats=Reference(ws,min_col=1,min_row=5,max_row=11)
ch.add_data(data,titles_from_data=True); ch.add_data(data2,titles_from_data=True); ch.set_categories(cats)
ch.height=11; ch.width=13
ws.add_chart(ch,'A14')
r=13
c=ws.cell(r,1,'Grafik di bawah akan terisi otomatis setelah kalian isi skornya.'); c.font=Font(italic=True,size=11,color='595959')

# ===== 2. HARI 2 DATA =====
ws=wb.create_sheet('H2 · DATA LAPANGAN')
head(ws,'HARI 2 · BENER — DATA LAPANGAN','Setiap kalimat HARUS punya bukti: angka, foto, atau kutipan. Tidak ada buktinya? CORET. Minimal 10 baris.',GOLD,7)
w(ws,[5,34,22,14,12,26,20])
th(ws,4,['No','Apa yang kami ukur / hitung','Di mana (sektor kami)','ANGKA','Satuan','Catatan / kejanggalan','Nama file foto'],GOLD)
for i in range(12):
    r=5+i; ws.cell(r,1,i+1).border=BD; ws.cell(r,1).alignment=Alignment('center'); blank(ws,r,r,2,7)
r=19
ws.cell(r,1,'HITUNG:').font=Font(bold=True,size=12,color=NAVY); r+=1
for lbl,f in [('TOTAL','=IFERROR(SUM(D5:D16),"")'),('RATA-RATA','=IFERROR(ROUND(AVERAGE(D5:D16),2),"")'),('TERBESAR','=IFERROR(MAX(D5:D16),"")'),('TERKECIL','=IFERROR(MIN(D5:D16),"")')]:
    ws.cell(r,2,lbl).font=Font(bold=True,size=11); ws.cell(r,2).border=BD; ws.cell(r,2).fill=PatternFill('solid',fgColor=FILL['GOLD'])
    c=ws.cell(r,3,f); c.border=BD; c.font=Font(bold=True,size=11); c.alignment=Alignment('center'); r+=1
r+=1
ws.cell(r,1,'SATU hal yang paling MENGEJUTKAN kami di lapangan:').font=Font(bold=True,size=12,color=NAVY); r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r+1,end_column=7); blank(ws,r,r+1,1,7)

# ===== 3. POHON AKAR =====
ws=wb.create_sheet('H2 · POHON AKAR MASALAH')
head(ws,'HARI 2 · BENER — POHON AKAR MASALAH (5x KENAPA)','Contoh: "Toilet bau" → kenapa? → "air tak jalan" → kenapa? → "keran rusak" → kenapa? → "tak ada yang lapor" → kenapa? → "tak tahu lapor ke mana". NAH ITU AKARNYA — dan itu bisa dikerjakan anak RPL.',GOLD,4)
w(ws,[28,50,30,20])
r=4
for lbl,ket in [('BUAH — akibat yang TERLIHAT & dirasakan orang',''),('BATANG — masalahnya apa',''),('AKAR 1 — penyebab SEBENARNYA',''),('AKAR 2',''),('AKAR 3','')]:
    c=ws.cell(r,1,lbl); c.font=Font(bold=True,size=11); c.fill=PatternFill('solid',fgColor=FILL['GOLD']); c.border=BD; c.alignment=Alignment(wrap_text=True,vertical='center')
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4); blank(ws,r,r,2,4); ws.row_dimensions[r].height=40; r+=1
r+=1
c=ws.cell(r,1,'AKAR YANG BAIK BISA DIKERJAKAN. AKAR YANG BURUK HANYA MENYALAHKAN ORANG.'); c.font=Font(bold=True,size=11,color='C00000'); r+=2
ws.cell(r,1,'RUMUSAN MASALAH KAMI:').font=Font(bold=True,size=12,color=NAVY); r+=1
for t in ['"Masalah sebenarnya adalah ______','karena ______','dibuktikan dengan data ______."']:
    c=ws.cell(r,1,t); c.font=Font(size=11,bold=True); c.border=BD; c.fill=PatternFill('solid',fgColor=FILL['GOLD'])
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4); blank(ws,r,r,2,4); ws.row_dimensions[r].height=30; r+=1

# ===== 4. HARI 3 WAWANCARA =====
ws=wb.create_sheet('H3 · WAWANCARA')
head(ws,'HARI 3 · BAGEUR — KUTIPAN LANGSUNG','Tulis kalimat mereka PERSIS, dalam tanda kutip. Jangan diringkas jadi kalimatmu sendiri. Minimal 3 narasumber.',ROSE,5)
w(ws,[5,24,20,50,22])
th(ws,4,['No','NAMA NARASUMBER','PERANNYA','KUTIPAN LANGSUNG ("...")','Di mana kami menemuinya'],ROSE)
for i in range(6):
    r=5+i; ws.cell(r,1,i+1).border=BD; ws.cell(r,1).alignment=Alignment('center'); blank(ws,r,r,2,5); ws.row_dimensions[r].height=34
r=13
ws.cell(r,1,'Kutipan mana yang paling MENUSUK perasaanmu? Kenapa?').font=Font(bold=True,size=12,color=NAVY); r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r+1,end_column=5); blank(ws,r,r+1,1,5)

# ===== 5. PETA EMPATI =====
ws=wb.create_sheet('H3 · PETA EMPATI')
head(ws,'HARI 3 · BAGEUR — PETA EMPATI','Pilih SATU narasumber utama. Isi 4 kuadran berdasarkan apa yang benar-benar kalian LIHAT & DENGAR — bukan tebakan.',ROSE,4)
w(ws,[42,42,10,10])
r=4
ws.cell(r,1,'NARASUMBER UTAMA:').font=Font(bold=True,size=11); ws.cell(r,1).border=BD; ws.cell(r,1).fill=PatternFill('solid',fgColor=FILL['ROSE'])
blank(ws,r,r,2,2); r+=2
q=[('APA YANG DIA KATAKAN\n(kutipan persisnya)','APA YANG DIA LAKUKAN\n(yang kalian lihat dengan mata sendiri)'),
   ('APA YANG DIA PIKIRKAN\n(dari nada bicara & ekspresinya)','APA YANG DIA RASAKAN\n(lelah? kesal? pasrah? bangga?)')]
for a,b in q:
    for i,t in enumerate([a,b],1):
        c=ws.cell(r,i,t); c.font=Font(bold=True,size=11,color='FFFFFF'); c.fill=PatternFill('solid',fgColor=ROSE)
        c.alignment=Alignment('center','center',wrap_text=True); c.border=BD
    ws.row_dimensions[r].height=40; r+=1
    ws.merge_cells(start_row=r,start_column=1,end_row=r+3,end_column=1)
    ws.merge_cells(start_row=r,start_column=2,end_row=r+3,end_column=2)
    blank(ws,r,r+3,1,2); r+=5

# ===== 6. MATRIKS PRIORITAS =====
ws=wb.create_sheet('H3 · MATRIKS PRIORITAS')
head(ws,'HARI 3 · BAGEUR — MATRIKS PRIORITAS & VOTING','Letakkan tiap masalah di kotak yang tepat. Yang masuk KANAN-ATAS = kandidat terkuat.',ROSE,4)
w(ws,[22,36,36,16])
r=4
th(ws,r,['','BISA KAMI KERJAKAN\n(2 hari, tanpa mesin)','SULIT KAMI KERJAKAN',''],ROSE); r+=1
for lbl,a,b,cl1,cl2 in [('DAMPAK BESAR\n(banyak orang terbantu)','KANDIDAT UTAMA — ambil yang ini','Butuh waktu lebih lama — persempit dulu','C8E6C9','FFF9C4'),
                        ('DAMPAK KECIL','Boleh, tapi kurang berarti','JANGAN DIAMBIL','FFE0B2','FFCDD2')]:
    c=ws.cell(r,1,lbl); c.font=Font(bold=True,size=11); c.fill=PatternFill('solid',fgColor=FILL['ROSE']); c.border=BD; c.alignment=Alignment('center','center',wrap_text=True)
    for i,(t,cl) in enumerate([(a,cl1),(b,cl2)],2):
        x=ws.cell(r,i,t); x.font=Font(bold=True,size=10,color='424242'); x.fill=PatternFill('solid',fgColor=cl); x.border=BD; x.alignment=Alignment('center','top',wrap_text=True)
    ws.row_dimensions[r].height=70; r+=1
r+=2
ws.cell(r,1,'HASIL VOTING KELAS — tulis besar-besar:').font=Font(bold=True,size=13,color=NAVY); r+=1
for lbl in ['MASALAH KAMI','ORANG YANG KAMI BANTU (sebut orangnya, bukan "semua orang")','TARGET SELESAI SELASA 21 JULI — bentuknya apa?']:
    c=ws.cell(r,1,lbl); c.font=Font(bold=True,size=11); c.fill=PatternFill('solid',fgColor=FILL['NAVY']); c.border=BD; c.alignment=Alignment(wrap_text=True,vertical='center')
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4); blank(ws,r,r,2,4); ws.row_dimensions[r].height=34; r+=1

# ===== 7. HARI 4 IDE & RANCANGAN =====
ws=wb.create_sheet('H4 · IDE & RANCANGAN')
head(ws,'HARI 4 · PINTER — IDE & RANCANGAN','JELEK DULUAN, BAGUS BELAKANGAN. Semua dikerjakan DI KELAS — tanpa bengkel, tanpa mesin, tanpa bahan kimia berbahaya.',NAVY,5)
w(ws,[5,40,26,26,20])
r=4
ws.cell(r,1,'BADAI IDE 6-3-5 — tulis semua ide, jangan disaring dulu').font=Font(bold=True,size=12,color=NAVY); r+=1
th(ws,r,['No','IDE','Siapa yang menulis','Dikembangkan oleh',''],BLUE); r+=1
for i in range(9):
    ws.cell(r,1,i+1).border=BD; ws.cell(r,1).alignment=Alignment('center'); blank(ws,r,r,2,5); r+=1
r+=1
ws.cell(r,1,'IDE YANG KAMI PILIH — harus lolos 4 SYARAT').font=Font(bold=True,size=12,color=NAVY); r+=1
for lbl in ['Ide kami','☐ 1. Memakai kompetensi keahlian kelas kami','☐ 2. BISA dikerjakan TANPA mesin & TANPA bahan berbahaya','☐ 3. Selesai dalam 2 hari','☐ 4. Benar-benar menolong orang yang kami wawancarai']:
    c=ws.cell(r,1,lbl); c.font=Font(bold=True,size=11); c.fill=PatternFill('solid',fgColor=FILL['NAVY']); c.border=BD; c.alignment=Alignment(wrap_text=True,vertical='center')
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5); blank(ws,r,r,2,5); ws.row_dimensions[r].height=28; r+=1
r+=1
c=ws.cell(r,1,'Ada satu saja yang tidak tercentang? BUANG IDENYA, pilih yang lain.'); c.font=Font(bold=True,size=11,color='C00000'); r+=2
ws.cell(r,1,'RANCANGAN TEKNIS (sesuai jurusan) — boleh foto sketsa tangan lalu tempel di sini').font=Font(bold=True,size=12,color=NAVY); r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r+7,end_column=5); blank(ws,r,r+7,1,5); r+=9
ws.cell(r,1,'DAFTAR BAHAN (utamakan KARDUS & BARANG BEKAS)').font=Font(bold=True,size=12,color=NAVY); r+=1
th(ws,r,['No','Bahan / Alat','Jumlah','Dari mana','Sudah dapat?'],BLUE); r+=1
for i in range(6):
    ws.cell(r,1,i+1).border=BD; ws.cell(r,1).alignment=Alignment('center'); blank(ws,r,r,2,5); r+=1

# ===== 8. SILIH ASAH =====
ws=wb.create_sheet('H4 · SILIH ASAH')
head(ws,'HARI 4 · PINTER — LEMBAR SILIH ASAH (umpan balik untuk tim lain)','DILARANG menulis "bagus" tanpa alasan. DILARANG menyerang orangnya — kritik idenya. Kritik yang jujur adalah bentuk kasih sayang.',NAVY,4)
w(ws,[24,30,30,32])
th(ws,4,['Tim yang kami nilai','Yang SUDAH JALAN','Yang BELUM JALAN','SATU saran dari kami'],BLUE)
for i in range(3):
    r=5+i; blank(ws,r,r,1,4); ws.row_dimensions[r].height=46
r=10
ws.cell(r,1,'UMPAN BALIK YANG KAMI TERIMA — dan SATU perbaikan yang akan kami kerjakan besok pagi:').font=Font(bold=True,size=12,color=NAVY); r+=1
th(ws,r,['Masukan yang paling menohok','SATU perbaikan yang akan kami kerjakan','',''],BLUE); r+=1
ws.merge_cells(start_row=r,start_column=2,end_row=r+2,end_column=2)
ws.merge_cells(start_row=r,start_column=3,end_row=r+2,end_column=4)
blank(ws,r,r+2,1,4)

# ===== 9. HARI 5 AKSI =====
ws=wb.create_sheet('H5 · AKSI & GELAR KARYA')
head(ws,'HARI 5 · SINGER — BUKTI AKSI NYATA','Sesuatu harus BENAR-BENAR BERUBAH di sekolah ini. Bukan dipamerkan — BERUBAH.',ORG,4)
w(ws,[32,40,32,20])
r=4
for lbl in ['Lokasi titik aksi kami','Apa PERSISNYA yang kami pasang / jalankan / perbaiki','Berapa orang yang akan terbantu? (angka, bukan "banyak")','Nama file foto BEFORE','Nama file foto AFTER','Kata MITRA kami setelah melihat hasilnya ("...")']:
    c=ws.cell(r,1,lbl); c.font=Font(bold=True,size=11); c.fill=PatternFill('solid',fgColor=FILL['ORG']); c.border=BD; c.alignment=Alignment(wrap_text=True,vertical='center')
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4); blank(ws,r,r,2,4); ws.row_dimensions[r].height=32; r+=1
r+=1
ws.cell(r,1,'PRESENTASI 60 DETIK — hafalkan formula ini untuk Gelar Karya').font=Font(bold=True,size=13,color=NAVY); r+=1
for lbl,ket in [('Kami MENEMUKAN','(sebut 1 ANGKA)'),('Kami MENDENGAR','(sebut 1 KUTIPAN narasumber)'),('Kami MEMBUAT','(sebut KARYA-nya)'),('Sekarang','(sebut PERUBAHAN-nya)')]:
    c=ws.cell(r,1,lbl); c.font=Font(bold=True,size=12,color=ORG); c.fill=PatternFill('solid',fgColor=FILL['ORG']); c.border=BD
    c2=ws.cell(r,2,ket); c2.font=Font(italic=True,size=10,color='595959'); c2.border=BD
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=4); blank(ws,r,r,3,4); ws.row_dimensions[r].height=30; r+=1

# ===== 10. RADAR ULANG =====
ws=wb.create_sheet('H5 · RADAR ULANG')
head(ws,'HARI 5 · SINGER — RADAR KEBIASAAN, MINGGU INI','Isi ULANG skormu. Bandingkan dengan Hari 1. Jujur — kalau tidak berubah, tulis tidak berubah.',ORG,4)
w(ws,[26,16,16,48])
th(ws,4,['KEBIASAAN','SKOR HARI 1','SKOR HARI 5','Apa yang berubah / kenapa tidak berubah?'],ORG)
r=5
for k,v in KEB:
    ws.cell(r,1,k).font=Font(bold=True,size=11); ws.cell(r,1).border=BD; ws.cell(r,1).fill=PatternFill('solid',fgColor=FILL['ORG'])
    blank(ws,r,r,2,4); r+=1
r+=1
c=ws.cell(r,1,'Lima hari terlalu singkat untuk mengubah kebiasaan. Yang berubah biasanya bukan kebiasaannya — tapi KESADARANNYA. Itu sudah cukup untuk memulai.')
c.font=Font(italic=True,size=11,color='595959')
r+=2
c=ws.cell(r,1,'JURNAL REFLEKSI & KONTRAK KEBIASAAN → DITULIS TANGAN DI BUKU TULISMU. Bukan di sini.')
c.font=Font(bold=True,size=12,color='C00000')

wb.save(f'{OUT}/TEMPLATE_DIGITAL_TIM.xlsx')
print('SAVED. sheets:',len(wb.sheetnames))
