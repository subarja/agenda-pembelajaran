#!/usr/bin/env python3
"""LOGISTIK_DAN_FORM.xlsx — paket bagikan, daftar ATK, spesifikasi form."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/sessions/vibrant-inspiring-einstein/mnt/Ko Kurikuler/SAKOLA_WALUYA_XI_2026/03_UNTUK_KOORDINATOR/LOGISTIK_DAN_FORM.xlsx"

NAVY = '1F3864'; HIJAU = '1FA971'; KUNING = 'F0B429'; KORAL = 'E5484D'; UNGU = '9B7BD4'; BIRU = '2E9BD6'
L_HIJAU = 'E4F5EC'; L_KUNING = 'FFF6E0'; L_BIRU = 'E6F2FB'; L_ABU = 'F2F5F6'; L_KORAL = 'FCE9E8'

thin = Side(style='thin', color='BFC9CC')
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()


def head(ws, row, cols, fill=NAVY):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = Font(bold=True, color='FFFFFF', size=10.5)
        cell.fill = PatternFill('solid', fgColor=fill)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BOX
    ws.row_dimensions[row].height = 32


def title(ws, text, sub, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(bold=True, size=14, color=NAVY)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    c = ws.cell(row=2, column=1, value=sub)
    c.font = Font(italic=True, size=10, color='5A6E72')
    ws.row_dimensions[1].height = 22


def rows(ws, start, data, fills=None, wrap_cols=()):
    r = start
    for i, d in enumerate(data):
        for j, v in enumerate(d, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = BOX
            c.alignment = Alignment(vertical='center', wrap_text=(j in wrap_cols))
            c.font = Font(size=10)
            if fills:
                f = fills(i, d)
                if f:
                    c.fill = PatternFill('solid', fgColor=f)
        r += 1
    return r


def widths(ws, w):
    for i, x in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = x


# ══════════════ 1. PAKET BAGIKAN ══════════════
ws = wb.active; ws.title = '1. PAKET BAGIKAN'
title(ws, 'APA YANG DIBAGIKAN, KE SIAPA, KAPAN, LEWAT APA',
      'Kolom "BOLEH DIBAGIKAN" = aman disebar. "JANGAN DIBAGIKAN" = dokumen internal, bukan untuk murid/orang tua.', 6)
head(ws, 4, ['PENERIMA', 'BERKAS', 'ISI SINGKAT', 'DIBAGIKAN LEWAT', 'KAPAN', 'STATUS'])

DATA = [
    # ---- WALI KELAS ----
    ('WALI KELAS\n(16 orang)', '01_UNTUK_GURU / KARTU_SAKU_GURU.docx',
     'Pegangan utama. 7 hlm A4, 1 hlm = 1 hari.', 'CETAK 1 set per wali kelas', 'H-2 (saat briefing)', 'BOLEH DIBAGIKAN'),
    ('WALI KELAS', '01_UNTUK_GURU / INSTRUMEN_ASESMEN.xlsx',
     'Rubrik + daftar nilai 36 murid. Deskripsi rapor terisi otomatis.', 'File Excel / Drive (1 salinan per kelas)', 'H-2', 'BOLEH DIBAGIKAN'),
    ('WALI KELAS', '01_UNTUK_GURU / PANDUAN_LENGKAP_GURU_rujukan.docx',
     'Rujukan lengkap. TIDAK wajib dibaca, TIDAK perlu dicetak massal.', 'Drive (baca saja) · cetak 2 eksemplar untuk ruang guru', 'H-2', 'BOLEH DIBAGIKAN'),
    ('WALI KELAS', '04_PAPARAN_GURU / 00_BRIEFING_GURU.pptx',
     'Materi briefing 60 menit.', 'Ditayangkan koordinator (tidak dibagikan filenya)', 'H-2', 'BOLEH DIBAGIKAN'),
    ('WALI KELAS', '02_UNTUK_MURID / PAPARAN_TAYANG/ (5 dek)',
     'Dek yang GURU PROYEKSIKAN ke murid. Guru tinggal klik next.', 'Salin ke laptop kelas / Drive', 'H-1 (sudah harus ada di laptop)', 'BOLEH DIBAGIKAN'),
    # ---- MURID ----
    ('MURID', 'Tautan GOOGLE SHEET TIM (salinan per tim)',
     'Tempat semua pekerjaan tim ditulis. 96 salinan (16 kelas x 6 tim).', 'Tautan di grup WA kelas + QR di layar', 'Hari 1 pagi', 'BOLEH DIBAGIKAN'),
    ('MURID', 'QR / tautan GOOGLE FORM "Suara Waluya"',
     'Formulir anonim. Diisi Hari 1.', 'QR ditayangkan di layar', 'Hari 1, sesi 09.30', 'BOLEH DIBAGIKAN'),
    ('MURID', '02_UNTUK_MURID / PAPAN_INSTRUKSI_HARIAN.docx',
     'Instruksi harian versi cetak. DITEMPEL DI DINDING, bukan dibagi per murid.', 'CETAK 1 set per kelas (16 set)', 'H-1', 'BOLEH DIBAGIKAN'),
    ('MURID', '05_GAMBAR / 17_Lembar_Cetak_Kartu_Tim.png',
     '6 kartu tim per kelas. Digunting, dilaminating.', 'CETAK 1 lembar per kelas (16 lembar)', 'H-2', 'BOLEH DIBAGIKAN'),
    ('MURID', '05_GAMBAR / poster Zona Z1-Z5 (A3)',
     '5 poster zona aksi, ditempel di kelas.', 'CETAK A3, 5 lembar per kelas (80 lembar)', 'H-1', 'BOLEH DIBAGIKAN'),
    ('MURID', '05_GAMBAR / 04_Template_Pohon_Akar_Masalah.png',
     'Ditayangkan / ditempel saat Hari 2.', 'Tayang di layar (sudah ada di dek Hari 2)', 'Hari 2', 'BOLEH DIBAGIKAN'),
    # ---- ORANG TUA ----
    ('ORANG TUA', 'Kontrak Kebiasaan (ditulis murid di buku tulis)',
     'Bukan file. Murid menulis sendiri, orang tua menandatangani di buku.', 'Dibawa pulang murid', 'Hari 5', 'BOLEH DIBAGIKAN'),
    # ---- INTERNAL ----
    ('KOORDINATOR\n(internal)', '03_UNTUK_KOORDINATOR / MATRIKS_OPERASIONAL_16_KELAS.xlsx',
     'Jadwal 16 kelas, sektor, gelombang, marshal, ceklis.', 'Pegangan koordinator + wakasek', 'H-5', 'JANGAN DIBAGIKAN ke murid'),
    ('KOORDINATOR\n(internal)', '03_UNTUK_KOORDINATOR / RANCANGAN_KOKURIKULER_resmi.docx',
     'Dokumen resmi/legal. Untuk kepala sekolah & pengawas.', 'Arsip sekolah', 'H-7', 'JANGAN DIBAGIKAN ke murid'),
    ('KOORDINATOR\n(internal)', '03_UNTUK_KOORDINATOR / DAFTAR_KEGIATAN_KOKURIKULER.xlsx',
     'Untuk diinput ke aplikasi rapor.', 'Operator rapor', 'Akhir semester', 'JANGAN DIBAGIKAN ke murid'),
    ('KOORDINATOR\n(internal)', 'Hasil Google Form "Suara Waluya" (spreadsheet)',
     'BERISI DATA SENSITIF. Kolom "ingin bicara dengan BK" HANYA untuk Guru BK.',
     'Akses dibatasi: Koordinator + Guru BK saja', 'Hari 1 sore', 'RAHASIA - jangan dibagikan ke wali kelas'),
    ('KOORDINATOR\n(internal)', '06_DRAFT_MD/ · 07_KODE/ · _ARSIP/',
     'Berkas kerja & arsip. Tidak dipakai saat kegiatan.', '-', '-', 'JANGAN DIBAGIKAN'),
]


def f_paket(i, d):
    s = d[5]
    if s.startswith('RAHASIA'): return L_KORAL
    if s.startswith('JANGAN'): return L_ABU
    if d[0].startswith('MURID'): return L_KUNING
    if d[0].startswith('ORANG'): return L_BIRU
    return L_HIJAU


end = rows(ws, 5, DATA, fills=f_paket, wrap_cols=(1, 2, 3, 4, 5, 6))
widths(ws, [16, 42, 46, 34, 20, 30])
ws.freeze_panes = 'A5'
for r in range(5, end):
    ws.row_dimensions[r].height = 42

ws.cell(row=end + 1, column=1, value='CATATAN PENTING').font = Font(bold=True, color=KORAL, size=11)
notes = [
    'MURID TIDAK MENERIMA LEMBAR KERJA CETAK. Yang dicetak hanya: Kartu Tim, Papan Instruksi Harian, dan poster zona - semuanya untuk KELAS, bukan per murid.',
    'Jurnal Refleksi (harian) dan Kontrak Kebiasaan (Hari 5) ditulis murid di BUKU TULIS mereka sendiri.',
    'Hasil Google Form "Suara Waluya" berisi jawaban tentang rasa tidak aman. Wali kelas TIDAK boleh melihat siapa menulis apa - hanya rekap tanpa nama yang ditayangkan.',
]
r = end + 2
for n in notes:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(row=r, column=1, value='- ' + n)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.font = Font(size=10)
    ws.row_dimensions[r].height = 30
    r += 1

# ══════════════ 2. DAFTAR ATK ══════════════
ws = wb.create_sheet('2. DAFTAR ATK')
title(ws, 'DAFTAR ATK & BAHAN — SAKOLA WALUYA',
      'Asumsi: 1 kelas = 36 murid = 6 tim.  16 kelas = +-576 murid = 96 tim.  Kolom terakhir untuk dicentang saat barang datang.', 7)
head(ws, 4, ['No', 'BARANG', 'SPESIFIKASI', 'PER 1 KELAS', 'TOTAL 16 KELAS', 'DIPAKAI KAPAN / UNTUK APA', 'ADA?'])

ATK = [
    # (barang, spek, per kelas, total, kegunaan)
    ('Kertas plano / flipchart', 'Ukuran +-65x100 cm, polos', '8 lembar', '128 lembar',
     'H1: norma tim (1). H2: grafik data per tim (6). H3: matriks prioritas (1). Cadangan sudah termasuk.'),
    ('Spidol whiteboard', 'Hitam + biru + merah', '4 buah', '64 buah', 'Setiap hari, di papan tulis.'),
    ('Spidol permanen / marker warna', 'Set 4 warna, ujung tebal', '6 set (1 per tim)', '96 set',
     'H2: menulis di plano. H3: peta empati. H4: menandai purwarupa.'),
    ('Sticky note / kertas tempel', '2 warna, 76x76 mm', '2 blok', '32 blok', 'H2: mengelompokkan temuan. H3: peta empati.'),
    ('Stiker titik / dot voting', 'Bulat warna, 3 per murid', '110 titik', '1.760 titik',
     'H3: voting prioritas masalah. Boleh diganti spidol (beri titik).'),
    ('Lakban kertas (masking tape)', '24 mm', '2 roll', '32 roll', 'Menempel plano & poster tanpa merusak dinding.'),
    ('Double tape busa', 'Tebal 1-3 mm, kuat', '2 roll', '32 roll', 'H5: memasang karya TANPA bor.'),
    ('Cable ties / kabel ties', '20-30 cm, campur ukuran', '1 pak (100 pcs)', '16 pak', 'H5: mengikat karya ke pagar/teralis/tiang.'),
    ('Lem tembak + isi (opsional)', 'Glue gun kecil + refill', '1 set', '16 set', 'H4: merakit purwarupa kardus. Diawasi Tim K3.'),
    ('Gunting', 'Ukuran sedang', '6 buah (1 per tim)', '96 buah', 'H4-H5: memotong kardus/kertas.'),
    ('Cutter + alas potong', 'Cutter kecil + alas kardus tebal', '3 set', '48 set', 'H4. HANYA dipakai murid, diawasi Tim K3.'),
    ('Penggaris besi 30-50 cm', '-', '3 buah', '48 buah', 'H4: memotong lurus & mengukur purwarupa.'),
    ('Meteran gulung 3-5 m', '-', '2 buah', '32 buah', 'H2: mengukur objek di sektor (panjang, tinggi, jarak).'),
    ('Timbangan dapur sederhana', 'Kapasitas 5 kg', '1 buah', '16 buah',
     'H2: menimbang sisa makanan MBG (food waste). BOLEH DIPAKAI BERGANTIAN - cukup 6-8 unit untuk 16 kelas.'),
    ('Sarung tangan karet', 'Sekali/berulang pakai', '12 pasang', '192 pasang', 'H2: mendata area kotor (toilet, tempat sampah).'),
    ('Masker', 'Masker medis biasa', '1 kotak (50)', '16 kotak', 'H2: area kotor. H4: mengamplas/memotong.'),
    ('Kardus bekas', 'Tebal, ukuran besar', '10 lembar', '160 lembar', 'H4: bahan utama purwarupa. GRATIS - kumpulkan dari kantin/koperasi/gudang sejak H-7.'),
    ('Pipa PVC bekas', '1/2 - 1 inci, potongan', 'seadanya', 'seadanya', 'H4: rangka purwarupa. Minta sisa proyek ke Sarpras.'),
    ('Botol & kemasan bekas', '-', 'seadanya', 'seadanya', 'H4: bahan purwarupa. Dikumpulkan murid sendiri.'),
    ('Kertas HVS bekas (1 sisi)', 'A4', '1 rim', '16 rim', 'H4: sketsa & gambar kerja. Pakai kertas bekas, jangan beli baru.'),
    ('Kertas HVS baru', 'A4 80 gsm', '-', '2 rim', 'Cetak lembar rubrik & administrasi koordinator.'),
    ('Kartu Tim (dicetak)', 'Dari 17_Lembar_Cetak_Kartu_Tim.png, dilaminating', '1 lembar (6 kartu)', '16 lembar',
     'Dipakai 5 hari. Laminating supaya awet.'),
    ('Papan Kartu', 'Karton tebal / sterofoam / papan kecil, 6 slot berlabel', '1 buah', '16 buah',
     'Ditempel di depan kelas. Bisa dibuat murid sendiri Hari 1.'),
    ('Poster Zona A3', '5 poster (Z1-Z5), cetak warna A3', '5 lembar', '80 lembar', 'H1: memilih Zona Aksi. Ditempel di dinding kelas.'),
    ('Papan Instruksi Harian', 'Cetak A4/A3, 6 halaman', '1 set', '16 set', 'Ditempel di kelas. Cadangan kalau proyektor mati.'),
    ('Kartu Saku Guru', 'Cetak 7 halaman A4', '1 set', '16 set', 'Pegangan wali kelas. Boleh dilaminating.'),
    ('Rompi / pita MARSHAL', 'Rompi OSIS atau pita lengan mencolok', '-', '12 buah',
     'Untuk 12 murid OSIS penjaga titik rawan (bukan per kelas).'),
    ('Kotak P3K', 'Isi standar', '-', '4 kotak', 'Di 4 titik POS kawasan. Bukan per kelas.'),
    ('Stopwatch', '-', '0 (pakai HP murid)', '0', 'Tim Waktu memakai HP sendiri. TIDAK perlu beli.'),
    ('Proyektor + kabel', 'Sudah ada di kelas', '1 unit', '16 unit', 'Menayangkan dek harian. PASTIKAN BERFUNGSI - cek H-1.'),
]

data = [(i + 1, a, b, c, d, e, '') for i, (a, b, c, d, e) in enumerate(ATK)]
end = rows(ws, 5, data,
           fills=lambda i, d: L_HIJAU if i % 2 == 0 else 'FFFFFF',
           wrap_cols=(2, 3, 4, 5, 6, 7))
widths(ws, [5, 26, 28, 18, 16, 52, 8])
ws.freeze_panes = 'A5'
for r in range(5, end):
    ws.row_dimensions[r].height = 34
for r in range(5, end):
    for col in (4, 5):
        ws.cell(row=r, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.cell(row=r, column=7).alignment = Alignment(horizontal='center', vertical='center')

r = end + 2
ws.cell(row=r, column=1, value='PRINSIP BELANJA').font = Font(bold=True, color=HIJAU, size=11)
for n in [
    'GRATIS DULU: kardus, botol, pipa PVC, kertas bekas - kumpulkan dari kantin, koperasi, gudang, dan rumah murid sejak H-7. Ini bagian dari pembelajaran, bukan kekurangan.',
    'PAKAI BERGANTIAN: timbangan, meteran, lem tembak tidak perlu 16 unit. Cukup 6-8 unit yang dijadwalkan bergantian antarkelas (lihat gelombang di Matriks Operasional).',
    'JANGAN BELI: stopwatch (pakai HP), alat bengkel, bor listrik - dilarang dipakai.',
    'YANG PALING KRITIS: kertas plano, spidol warna, sarung tangan, masker, cable ties, double tape. Tanpa ini kegiatan H2/H4/H5 macet.',
]:
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    c = ws.cell(row=r, column=1, value='- ' + n)
    c.alignment = Alignment(wrap_text=True, vertical='top'); c.font = Font(size=10)
    ws.row_dimensions[r].height = 30

# ══════════════ 3. GOOGLE FORM ══════════════
ws = wb.create_sheet('3. GOOGLE FORM')
title(ws, 'GOOGLE FORM "SUARA WALUYA" — daftar pertanyaan lengkap',
      'Dibuat Tim ICT. SATU form untuk 16 kelas. Diisi Hari 1 (Rabu 15 Juli), sesi 09.30, +-12 menit.', 5)

ws.cell(row=4, column=1, value='PENGATURAN FORM (WAJIB, sebelum dibagikan)').font = Font(bold=True, size=11, color=KORAL)
SET = [
    ('Kumpulkan alamat email', 'MATI / OFF', 'Kalau menyala, form TIDAK anonim lagi. Ini yang paling sering terlewat.'),
    ('Batasi 1 respons per orang', 'MATI / OFF', 'Membutuhkan login = tidak anonim.'),
    ('Judul form', 'SUARA WALUYA', '-'),
    ('Deskripsi form', 'Formulir ini ANONIM. Gurumu tidak tahu siapa menulis apa. Jawablah jujur - tidak ada jawaban yang salah, dan tidak ada yang dihukum.', 'Tampilkan apa adanya.'),
    ('Akses hasil (spreadsheet)', 'Koordinator + Guru BK SAJA', 'Wali kelas TIDAK diberi akses. Mereka hanya melihat rekap tanpa nama.'),
    ('Tanggapan diterima', 'Tutup form Hari 1 pukul 10.30', 'Supaya rekap bisa langsung ditayangkan di sesi berikutnya.'),
]
head(ws, 5, ['PENGATURAN', 'HARUS DISET', 'KENAPA'])
end = rows(ws, 6, SET, fills=lambda i, d: L_KORAL, wrap_cols=(1, 2, 3))

r = end + 2
ws.cell(row=r, column=1, value='DAFTAR PERTANYAAN').font = Font(bold=True, size=11, color=NAVY)
head(ws, r + 1, ['No', 'PERTANYAAN (salin apa adanya)', 'JENIS', 'WAJIB?', 'CATATAN'])
Q = [
    (1, 'Kelas kamu apa?', 'Pilihan ganda (daftar 16 kelas XI)', 'Wajib',
     'Ini SATU-SATUNYA penanda. Tidak menunjuk orang, hanya kelas - supaya rekap bisa dipisah per kelas.'),
    (2, 'Sebutkan SATU hal yang membuat kamu BETAH di sekolah ini.', 'Jawaban panjang', 'Wajib',
     'Sengaja ditanya duluan supaya murid tidak langsung defensif.'),
    (3, 'Sebutkan SATU hal yang membuat kamu TIDAK NYAMAN di sekolah ini.', 'Jawaban panjang', 'Wajib',
     'Inti data Hari 1. Jawaban inilah yang dikelompokkan di papan tulis.'),
    (4, 'Sebutkan SATU hal yang membuat kamu merasa TIDAK AMAN di sekolah ini.', 'Jawaban panjang', 'TIDAK wajib',
     'BOLEH DIKOSONGKAN. Jangan dipaksa. Rapor D.4.4 menunjukkan perundungan meningkat - pertanyaan ini penting tapi harus sukarela.'),
    (5, 'Seberapa AMAN kamu merasa di sekolah ini?', 'Skala linear 1-5\n(1 = sangat tidak aman, 5 = sangat aman)', 'Wajib',
     'Angka. Dipakai sebagai data pembanding di akhir semester.'),
    (6, 'Seberapa BETAH kamu belajar di sekolah ini?', 'Skala linear 1-5', 'Wajib',
     'Terkait indikator D.4.1 (wellbeing) dan D.19.5 (gemar belajar).'),
    (7, 'Menurutmu, masalah mana yang PALING MENDESAK diperbaiki di sekolah ini?',
     'Pilihan ganda:\n- Kebersihan & sanitasi (toilet, wastafel, air)\n- Makanan & gizi (kantin, MBG, sampah makanan)\n- Kenyamanan belajar (kelas, ventilasi, kebisingan)\n- Kesehatan & kebugaran\n- Rasa aman & keramahan (perundungan, sikap)',
     'Wajib', 'Kelima pilihan ini = 5 Zona Aksi. Hasilnya dipakai kelas untuk memilih zona.'),
    (8, 'Apakah kamu ingin berbicara langsung dengan Guru BK?', 'Kotak centang (checkbox), 1 pilihan:\n[ ] Ya, saya ingin bicara dengan Guru BK',
     'TIDAK wajib', 'PENTING: hasil kolom ini HANYA dilihat Guru BK. Jangan pernah ditayangkan di kelas.'),
    (9, 'Kalau kamu mencentang di atas, tulis namamu di sini (boleh dikosongkan).', 'Jawaban singkat', 'TIDAK wajib',
     'Satu-satunya kolom nama. Sukarela. Kalau dikosongkan, Guru BK tetap tahu ada murid di kelas itu yang butuh bantuan.'),
]
end = rows(ws, r + 2, Q, fills=lambda i, d: L_KORAL if d[0] in (8, 9) else (L_HIJAU if i % 2 == 0 else 'FFFFFF'),
           wrap_cols=(2, 3, 5))
widths(ws, [5, 52, 34, 12, 58])
for rr in range(r + 2, end):
    ws.row_dimensions[rr].height = 52

r = end + 2
ws.cell(row=r, column=1, value='YANG DITAYANGKAN KE KELAS').font = Font(bold=True, size=11, color=HIJAU)
for n in [
    'HANYA hasil pertanyaan 2, 3, 4, dan 7 - dalam bentuk daftar jawaban TANPA NAMA dan TANPA KELAS LAIN.',
    'Pertanyaan 8 dan 9 TIDAK PERNAH ditayangkan. Koordinator meneruskan langsung ke Guru BK di hari yang sama.',
    'Cara menayangkan: buka spreadsheet hasil, filter kolom "Kelas" = kelas itu, sembunyikan kolom 8-9, lalu tayangkan.',
]:
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value='- ' + n)
    c.alignment = Alignment(wrap_text=True, vertical='top'); c.font = Font(size=10)
    ws.row_dimensions[r].height = 28

# ══════════════ 4. GOOGLE SHEET TIM ══════════════
ws = wb.create_sheet('4. GOOGLE SHEET TIM')
title(ws, 'GOOGLE SHEET TIM — struktur & isi tiap lembar',
      'Template: 02_UNTUK_MURID/TEMPLATE_DIGITAL_TIM.xlsx. Unggah ke Drive, lalu SALIN 96x (16 kelas x 6 tim). Nama file: "XI TKR 1 - TIM 3".', 5)
head(ws, 4, ['LEMBAR', 'DIISI KAPAN', 'KOLOM / PERTANYAAN YANG HARUS ADA', 'DIISI SIAPA', 'DIPAKAI UNTUK'])

SH = [
    ('H1 - RADAR KEBIASAAN', 'Hari 1, 25 menit',
     'Nama murid | Bangun pagi (1-10) | Beribadah | Berolahraga | Makan sehat | Gemar belajar | Bermasyarakat | Tidur cepat | Rata-rata (rumus)',
     'Tiap murid mengisi barisnya sendiri', 'Grafik radar otomatis. Dibandingkan lagi di Hari 5.'),
    ('H2 - DATA LAPANGAN', 'Hari 2, 55 menit',
     'No | Objek yang diukur | Lokasi persis | ANGKA (wajib) | Satuan | Waktu ukur | Diukur siapa | Tautan foto',
     'Tim (Kapten membagi tugas)', 'WAJIB minimal 10 baris angka + 5 foto per tim.'),
    ('H2 - POHON AKAR', 'Hari 2, 40 menit',
     'BUAH (akibat terlihat) | BATANG (masalahnya) | AKAR 1 | AKAR 2 | AKAR 3 | Bukti untuk tiap akar (angka/foto/kutipan)',
     'Tim, berdiskusi', 'Kolom "bukti" tidak boleh kosong. Tanpa bukti = coret.'),
    ('H3 - KUTIPAN WAWANCARA', 'Hari 3, 60 menit',
     'Narasumber | Perannya | Pertanyaan | KUTIPAN PERSIS (tanda kutip) | Perasaan yang terdengar | Izin rekam? (Ya/Tidak)',
     'Tim (2-3 orang delegasi)', 'Kutipan harus PERSIS, bukan rangkuman.'),
    ('H3 - PETA EMPATI', 'Hari 3, 30 menit',
     'Siapa dia | Yang DIA KATAKAN | Yang DIA LAKUKAN | Yang DIA PIKIRKAN | Yang DIA RASAKAN | Yang MEMBUATNYA SUSAH | Yang DIA BUTUHKAN',
     'Tim', 'Hanya boleh diisi dari yang DILIHAT & DIDENGAR, bukan tebakan.'),
    ('H3 - MATRIKS PRIORITAS', 'Hari 3, 25 menit',
     'Masalah | Dampak (1-5) | Bisa kami kerjakan? (1-5) | Skor total (rumus) | Peringkat (rumus)',
     'Tim, lalu voting kelas', 'Menentukan SATU masalah yang dikerjakan tim itu.'),
    ('H4 - IDE & RANCANGAN', 'Hari 4, 90 menit',
     'Ide (min. 18 dari Badai Ide 6-3-5) | Ide terpilih | Kenapa dipilih | Bahan | Ukuran | Siapa mengerjakan apa | Sketsa (tautan foto)',
     'Tim', 'Ide harus memakai KOMPETENSI JURUSAN tim itu.'),
    ('H4 - SILIH ASAH', 'Hari 4, 30 menit',
     'Tim penilai | Yang SUDAH BAIK | Yang PERLU DIPERBAIKI | Satu saran konkret | Ditindaklanjuti? (Ya/Tidak/Alasan)',
     'Tim LAIN yang mengisi', 'Kritik idenya, jangan orangnya. Bagian dari nilai KOLABORASI.'),
    ('H5 - AKSI & BUKTI', 'Hari 5, 90 menit',
     'Titik pemasangan | Foto SEBELUM | Foto SESUDAH | Cara memasang (tanpa bor) | Berfungsi? | Siapa yang akan merawat | Kendala',
     'Tim (Tim Dokumentasi memimpin)', 'Foto before-after WAJIB. Ini bukti utama Gelar Karya.'),
    ('H5 - SERAH TERIMA', 'Hari 5, saat Gelar Karya',
     'Nama karya | Butuh fabrikasi mesin? | Gambar kerja (tautan) | Diserahkan kepada | Tanggal | Tanda terima',
     'Tim yang karyanya butuh mesin', 'Penyerahan resmi ke Wakasek Sarpras TETAP dihitung sebagai Aksi Nyata.'),
    ('CEKLIS TIM', 'Setiap hari',
     'Hari | Target hari ini | Selesai? | Siapa tidak hadir | Catatan Kapten',
     'Kapten Tim', 'Guru cukup melihat lembar ini untuk tahu tim mana yang tertinggal.'),
]
end = rows(ws, 5, SH, fills=lambda i, d: L_BIRU if i % 2 == 0 else 'FFFFFF', wrap_cols=(1, 2, 3, 4, 5))
widths(ws, [24, 18, 62, 24, 44])
ws.freeze_panes = 'A5'
for rr in range(5, end):
    ws.row_dimensions[rr].height = 48

r = end + 2
ws.cell(row=r, column=1, value='CARA MENYIAPKAN (Tim ICT, H-3)').font = Font(bold=True, size=11, color=BIRU)
for n in [
    '1. Unggah TEMPLATE_DIGITAL_TIM.xlsx ke Google Drive, buka sebagai Google Sheet.',
    '2. Buat folder per kelas. Salin template 6x di tiap folder. Ganti nama: "XI TKR 1 - TIM 1" dst. Total 96 salinan.',
    '3. Setel akses: "Siapa saja yang memiliki link" -> EDITOR. Jangan minta login - banyak murid tidak punya akun sekolah.',
    '4. Kumpulkan 96 tautan itu dalam satu spreadsheet daftar, bagikan ke 16 wali kelas (masing-masing hanya 6 tautan kelasnya).',
    '5. Wali kelas menyebarkan 6 tautan ke grup WA kelas pada Hari 1 pagi.',
]:
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value=n)
    c.alignment = Alignment(wrap_text=True, vertical='top'); c.font = Font(size=10)
    ws.row_dimensions[r].height = 26

wb.save(OUT)
print("SAVED:", OUT)
print("sheets:", wb.sheetnames)
