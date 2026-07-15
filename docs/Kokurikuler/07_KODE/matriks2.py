import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
OUT="/sessions/vibrant-inspiring-einstein/mnt/Ko Kurikuler/SAKOLA_WALUYA_XI_2026/03_UNTUK_KOORDINATOR"
F=f'{OUT}/MATRIKS_OPERASIONAL_16_KELAS.xlsx'
wb=openpyxl.load_workbook(F)
NAVY='1F3864'; BIRU='DEEBF7'; JINGGA='FCE4D6'; ROSE='FCE4EC'; GOLD='FFF2CC'; GRN='E2EFDA'
thin=Side(style='thin',color='B0BEC5'); BD=Border(left=thin,right=thin,top=thin,bottom=thin)
def hdr(ws,row,cols,fill=NAVY):
    for i,c in enumerate(cols,1):
        x=ws.cell(row=row,column=i,value=c); x.font=Font(bold=True,color='FFFFFF',size=10,name='Calibri')
        x.fill=PatternFill('solid',fgColor=fill); x.alignment=Alignment('center','center',wrap_text=True); x.border=BD
    ws.row_dimensions[row].height=34
def title(ws,t,s,n):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n)
    c=ws.cell(row=1,column=1,value=t); c.font=Font(bold=True,size=15,color=NAVY,name='Calibri'); ws.row_dimensions[1].height=24
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=n)
    c=ws.cell(row=2,column=1,value=s); c.font=Font(italic=True,size=10,color='595959',name='Calibri'); c.alignment=Alignment(wrap_text=True,vertical='top'); ws.row_dimensions[2].height=30
def w(ws,ws_):
    for i,x in enumerate(ws_,1): ws.column_dimensions[get_column_letter(i)].width=x
def put(ws,r,v,fill=None,h=32,bold=False):
    for i,x in enumerate(v,1):
        c=ws.cell(row=r,column=i,value=x); c.font=Font(size=10,name='Calibri',bold=bold)
        c.alignment=Alignment(vertical='center',wrap_text=True); c.border=BD
        if fill: c.fill=PatternFill('solid',fgColor=fill)
    ws.row_dimensions[r].height=h

# ===== 4. JADWAL NGAJUGJUG =====
ws=wb.create_sheet('4. JADWAL NGAJUGJUG')
title(ws,'JADWAL NGAJUGJUG — MENGHAMPIRI NARASUMBER (HARI 3, JUMAT 17 JULI)',
 'Narasumber TIDAK dikumpulkan di satu tempat. Mereka BEKERJA SEPERTI BIASA. MURID yang mendatangi mereka di tempat kerjanya. Beban tiap narasumber: hanya 2 kelas × 10 menit = 20 menit total.',7)
hdr(ws,4,['KODE','NARASUMBER','DI MANA DIA BIASANYA BERADA\n(koordinator WAJIB isi lokasi nyatanya)','KELAS PENGHAMPIR\n(Gelombang BIRU)','SLOT BIRU','KELAS PENGHAMPIR\n(Gelombang JINGGA)','SLOT JINGGA'])
w(ws,[7,30,40,26,14,26,14])
N=[
 ('N1','Petugas Kebersihan (Caraka) 1','..................................................\n(mis. area toilet Gedung A / gudang alat kebersihan)','XI MEKATRONIKA - A','08.15–08.25','XI KIMIA INDUSTRI - B','09.30–09.40'),
 ('N2','Petugas Kebersihan (Caraka) 2','..................................................','XI KIMIA INDUSTRI - A','08.15–08.25','XI TEKNIK PEMESINAN - A','09.30–09.40'),
 ('N3','Ibu / Bapak Kantin','..................................................\n(kantin sekolah)','XI DKV - A','08.15–08.25','XI RPL - A','09.30–09.40'),
 ('N4','Penanggung Jawab MBG / petugas SPPG','..................................................\n(titik distribusi MBG)','XI KIMIA INDUSTRI - C','08.15–08.25','XI DKV - B','09.30–09.40'),
 ('N5','Petugas TU / Perpustakaan','..................................................','XI MEKATRONIKA - D','08.15–08.25','XI RPL - B','09.30–09.40'),
 ('N6','Satpam','..................................................\n(pos satpam / gerbang)','XI ANIMASI - B','08.15–08.25','XI DKV - C','09.30–09.40'),
 ('N7','Pembina Ekskul / staf Kesiswaan','..................................................','XI ANIMASI - A','08.15–08.25','XI MEKATRONIKA - C','09.30–09.40'),
 ('N8','Tukang Kebun / Penjaga Parkir','..................................................','XI MEKATRONIKA - B','08.15–08.25','XI TEKNIK PEMESINAN - B','09.30–09.40'),
]
r=5
for x in N:
    put(ws,r,list(x),h=42)
    ws.cell(row=r,column=4).fill=PatternFill('solid',fgColor=BIRU)
    ws.cell(row=r,column=5).fill=PatternFill('solid',fgColor=BIRU)
    ws.cell(row=r,column=6).fill=PatternFill('solid',fgColor=JINGGA)
    ws.cell(row=r,column=7).fill=PatternFill('solid',fgColor=JINGGA)
    r+=1

r+=1
ws.cell(row=r,column=1,value='ATURAN NGAJUGJUG — WAJIB DIBACAKAN GURU SEBELUM MURID BERANGKAT').font=Font(bold=True,size=11,color='C00000',name='Calibri')
rules=[
 '1. YANG BOLEH MENGHAMPIRI HANYA DELEGASI KELAS: 3 murid saja (1 Kapten Tim + 1 pencatat + 1 dokumentasi). BUKAN seluruh tim, BUKAN seluruh kelas. Hasilnya dibagikan ke seluruh kelas setelah kembali.',
 '2. MAKSIMAL 10 MENIT. Tim Waktu kelas yang mengingatkan. Lewat 10 menit = pamit, apa pun kondisinya.',
 '3. NARASUMBER SEDANG BEKERJA. Murid MENGHAMPIRI, memberi salam, MENUNJUKKAN Kartu Temu, lalu MEMINTA IZIN: "Bapak/Ibu, boleh minta waktu 10 menit? Kalau sedang sibuk, kami bisa kembali nanti."',
 '4. KALAU NARASUMBER MENOLAK ATAU SEDANG SIBUK: JANGAN DIPAKSA. Catat "tidak tersedia" di Kartu Temu, kembali ke POS, lapor ke wali kelas. Wali kelas menghubungi Koordinator untuk penjadwalan ulang atau narasumber cadangan.',
 '5. KALAU NARASUMBER TIDAK DITEMUKAN DI TEMPATNYA: cari maksimal 5 menit, lalu kembali ke POS. JANGAN berkeliaran mencari. Lapor.',
 '6. TIDAK BOLEH MEREKAM tanpa izin lisan. Tanyakan dulu: "Boleh kami rekam suaranya, Pak/Bu?"',
 '7. PULANG MEMBAWA: minimal 3 KUTIPAN LANGSUNG (kalimat PERSIS narasumber, dalam tanda kutip) + nama & perannya.',
 '8. SEBELUM PAMIT: UCAPKAN TERIMA KASIH, dan katakan: "Hari Selasa kami akan tunjukkan hasilnya ke Bapak/Ibu." — Lalu TEPATI di Gelar Karya.',
 '9. TIM YANG TIDAK KEBAGIAN NARASUMBER KUNCI tetap mewawancarai warga sekolah biasa DI SEKTORNYA: murid kelas X yang sedang MPLS (hanya saat istirahat MPLS!), murid kelas lain, guru piket, petugas lain. Tidak perlu jadwal khusus.',
]
for x in rules:
    r+=1
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
    c=ws.cell(row=r,column=1,value=x); c.font=Font(size=10,name='Calibri'); c.alignment=Alignment(wrap_text=True,vertical='top')
    ws.row_dimensions[r].height=30
r+=2
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
c=ws.cell(row=r,column=1,value='H-1 (KAMIS 16 JULI): Koordinator MENEMUI setiap narasumber, memberi tahu mereka akan didatangi 2 kelas (10 menit masing-masing) pada Jumat, dan MENEMPELKAN KARTU JADWAL kecil di tempat kerjanya. Ini bukan formalitas — tanpa ini, narasumber akan merasa diserbu dan menolak.')
c.font=Font(size=10,bold=True,color='C00000',name='Calibri'); c.alignment=Alignment(wrap_text=True,vertical='top')
c.fill=PatternFill('solid',fgColor=ROSE); ws.row_dimensions[r].height=44
ws.freeze_panes='A5'
print('sheet4 ok')

# ===== 5. HARI 5 & GELAR KARYA =====
ws=wb.create_sheet('5. HARI 5 & GELAR KARYA')
title(ws,'JADWAL HARI 5 — SELASA, 21 JULI 2026 · SINGER · "BERTINDAK & BERBAGI"',
 'AKSI NYATA bergelombang (agar 16 kelas tidak memasang serentak). GELAR KARYA LORONG — stan di depan kelas masing-masing, tidak perlu aula.',4)
hdr(ws,4,['WAKTU','GELOMBANG BIRU (8 kelas)','GELOMBANG JINGGA (8 kelas)','KETERANGAN'])
w(ws,[18,50,50,40])
J5=[
 ('06.30–07.30','APEL + MBG BERSAMA + NGARIUNG','APEL + MBG + NGARIUNG','—'),
 ('07.30–07.45','PEMANTIK — Perbaikan Kilat (15 menit, guru diam)','PEMANTIK — Perbaikan Kilat','Menutup masukan Silih Asah dari Hari 4'),
 ('07.45–08.30\n(45 menit)','AKSI NYATA DI TITIK SASARAN (sektor sendiri)\n• Foto BEFORE (wajib)\n• Pasang / jalankan / uji\n• Foto AFTER (sudut sama)','DI KELAS — siapkan STAN LORONG + finalisasi karya + latihan presentasi 60 detik','Hanya 8 kelas bergerak. Pemasangan TANPA BOR: cable ties / double tape / freestanding'),
 ('08.30–09.15\n(45 menit)','DI KELAS — siapkan STAN LORONG + latihan presentasi 60 detik','AKSI NYATA DI TITIK SASARAN (sektor sendiri)\n• Foto BEFORE → pasang → Foto AFTER','Bertukar'),
 ('09.15–09.30','JEDA GERAK & SNACK','JEDA GERAK & SNACK','—'),
 ('09.30–09.45','TATA STAN di depan ruang kelas masing-masing','TATA STAN di depan ruang kelas masing-masing','Wajib pajang: karya · plano data & grafik · foto BEFORE–AFTER · kutipan narasumber · Peta Empati'),
 ('09.45–10.15\n(30 menit)','RONDE 1 — TUAN RUMAH (jaga stan, presentasi 60 detik ke setiap pengunjung)','RONDE 1 — BERKUNJUNG (rute searah, lihat tab 6)','Penilai berkeliling: Kepala Sekolah, Wakasek, Kaprog (menilai kelas jurusannya), OSIS'),
 ('10.15–10.45\n(30 menit)','RONDE 2 — BERKUNJUNG','RONDE 2 — TUAN RUMAH','Bertukar peran'),
 ('10.45–11.00\n(15 menit)','PENYERAHAN RESMI gambar kerja & surat permohonan pemasangan kepada Wakasek Sarpras / Kaprog + APRESIASI','PENYERAHAN RESMI + APRESIASI','Kategori apresiasi: "Data Paling Jujur" · "Karya Paling Berguna" · "Tim Paling Silih Asuh" · "Kegagalan Paling Berharga"'),
 ('11.00–11.45\n(45 menit)','KAMPANYE DIGITAL (serahkan konten ke Tim ICT) + RADAR KEBIASAAN ULANG + KONTRAK KEBIASAAN WALUYA','Sama','Kontrak dibawa pulang untuk TTD orang tua, dikembalikan paling lambat Jumat 24 Juli 2026'),
 ('11.45–12.00','REFLEKSI PENUTUP & JURNAL — kumpulkan seluruh LK','Sama','Portofolio asesmen (bobot 30%)'),
]
r=5
for x in J5:
    put(ws,r,list(x),h=(64 if '•' in x[1] or '•' in x[2] else 38))
    ws.cell(row=r,column=2).fill=PatternFill('solid',fgColor=BIRU)
    ws.cell(row=r,column=3).fill=PatternFill('solid',fgColor=JINGGA)
    r+=1
r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
c=ws.cell(row=r,column=1,value='PENILAI GELAR KARYA (tanpa guru tambahan): Wali kelas menilai timnya sendiri saat presentasi ke pengunjung · 6 Kaprog berkeliling menilai kelas jurusannya masing-masing (Mekatronika 4 kelas, DKV 3, Kimia 3, RPL 2, Animasi 2, Pemesinan 2) · Kepala Sekolah, Wakasek, Komite, OSIS berkeliling · MITRA YANG DIWAWANCARAI (Caraka, Satpam, Ibu Kantin) DIUNDANG — ini penting, mereka sudah dijanjikan.')
c.font=Font(size=10,bold=True,name='Calibri'); c.alignment=Alignment(wrap_text=True,vertical='top'); c.fill=PatternFill('solid',fgColor=GRN); ws.row_dimensions[r].height=48
ws.freeze_panes='A5'
print('sheet5 ok')

# ===== 6. MARSHAL & ZONA MERAH =====
ws=wb.create_sheet('6. MARSHAL & ZONA MERAH')
title(ws,'PENGAWASAN LAPANGAN — TANPA GURU TAMBAHAN','Koordinator sendirian tidak mungkin mengawasi 48 tim. Pengawasan dibagi ke titik simpul, dijaga OSIS ber-rompi + satpam.',4)
hdr(ws,4,['TITIK SIMPUL','SIAPA YANG MENJAGA','TUGAS','JUMLAH'])
w(ws,[34,30,58,10])
M=[
 ('Tangga utama (tiap lantai)','MARSHAL WALUYA — murid OSIS/MPK ber-rompi','Mengarahkan arus naik/turun. Memastikan tidak ada murid sendirian. Mencatat tim yang terlihat keluar sektornya.','4–6'),
 ('Koridor penghubung antargedung','MARSHAL WALUYA (OSIS)','Titik lapor cepat. Memegang daftar sektor tiap kelas — kalau ada tim di sektor yang salah, ingatkan & catat.','2–4'),
 ('Depan area MPLS kelas X (ZONA MERAH)','MARSHAL WALUYA + panitia MPLS','MENAHAN tim XI yang mencoba masuk. Kelas XI HANYA boleh masuk pada slot wawancara adik kelas X yang sudah dikoordinasikan.','2'),
 ('Gerbang & area parkir','SATPAM SEKOLAH','Memastikan TIDAK ADA MURID KELUAR GERBANG. Ini garis merah — tidak boleh dilanggar.','2–3'),
 ('Kantin & titik distribusi MBG','MARSHAL WALUYA (OSIS)','Mengatur antrean tim yang mengambil data. Hanya kelas ber-sektor kantin/MBG yang boleh mengukur di sini.','2'),
 ('Keliling seluruh sekolah','KOORDINATOR KOKURIKULER','Pengawas umum. Menerima laporan dari marshal. Menangani tim yang tersesat/telat. Menjadi tempat bertanya wali kelas yang kewalahan.','1'),
 ('Keliling seluruh sekolah','WAKASEK KESISWAAN','Pengawas umum kedua. Berkoordinasi dengan panitia MPLS.','1'),
 ('POS UKS','Petugas UKS','Siaga menerima murid cedera. Kotak P3K juga ada di tiap POS kelas.','1'),
]
r=5
for x in M:
    put(ws,r,list(x),h=42)
    if 'ZONA MERAH' in x[0] or 'Gerbang' in x[0]: 
        for i in range(1,5): ws.cell(row=r,column=i).fill=PatternFill('solid',fgColor=ROSE)
    r+=1
r+=2
ws.cell(row=r,column=1,value='ZONA MERAH — TIDAK BOLEH DIMASUKI TIM KELAS XI').font=Font(bold=True,size=12,color='C00000',name='Calibri')
zm=[
 '1. AREA MPLS KELAS X — kecuali pada slot wawancara adik kelas X yang sudah dikoordinasikan dengan panitia MPLS. Maksimal 2 kelas XI per slot, HANYA saat jam istirahat MPLS. Kelas X adalah narasumber yang sangat berharga (mereka "mata baru" yang melihat sekolah apa adanya) — tapi jangan sampai MPLS mereka terganggu.',
 '2. RUANG YANG SEDANG DIPAKAI KBM — kalau ada.',
 '3. RUANG GURU, RUANG KEPALA SEKOLAH, RUANG TU (kecuali slot ngajugjug ke petugas TU).',
 '4. BENGKEL & LABORATORIUM PRAKTIK — semua, tanpa kecuali. Tidak ada toolman yang mendampingi.',
 '5. DI LUAR GERBANG SEKOLAH — garis merah mutlak.',
 '6. ATAP, TANGGA DARURAT, RUANG PANEL LISTRIK, TANDON AIR bagian atas — semua yang menuntut naik ketinggian lebih dari 1 meter.',
]
for x in zm:
    r+=1
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
    c=ws.cell(row=r,column=1,value=x); c.font=Font(size=10,name='Calibri'); c.alignment=Alignment(wrap_text=True,vertical='top')
    c.fill=PatternFill('solid',fgColor=ROSE); c.border=BD
    ws.row_dimensions[r].height=(44 if len(x)>150 else 26)
ws.freeze_panes='A5'
print('sheet6 ok')

# ===== 7. CEKLIS KOORDINATOR H-3 =====
ws=wb.create_sheet('7. CEKLIS KOORDINATOR')
title(ws,'CEKLIS KOORDINATOR — KHUSUS SKALA 16 KELAS','Selesai paling lambat Minggu, 12 Juli 2026. Tanpa ini, hari pertama akan kacau.',4)
hdr(ws,4,['✓','ITEM','KAPAN','PENANGGUNG JAWAB'])
w(ws,[5,74,20,26])
C=[
 ('','PETAKAN 16 SEKTOR ke DENAH NYATA SMKN 2 Cimahi. Ganti "GEDUNG/BLOK A–D" di tab 1 dengan nama gedung & lantai yang sebenarnya. Cetak PETA SEKTOR untuk tiap kelas.','H-5','Koordinator + Wakasek Sarpras'),
 ('','Cetak KARTU SEKTOR per kelas (16 lembar A4): nama kelas, gelombang, zona, daftar objek EKSKLUSIF, batas wilayah, dan apa yang TIDAK boleh disentuh.','H-4','Koordinator'),
 ('','Siapkan PAPAN KARTU: 6 SLOT + 6 KARTU TIM (warna berbeda) per kelas × 16 kelas. Cetak lembar 17_Lembar_Cetak_Kartu_Tim.png (folder 05_GAMBAR), gunting, laminating/tempel di karton. Murah, dipakai berulang. GURU TIDAK MENANDATANGANI APA PUN.','H-4','Koordinator + TU'),
 ('','Cetak PAPAN INSTRUKSI HARIAN (02_UNTUK_MURID) — 1 set per kelas untuk ditayangkan/ditempel.','H-4','Koordinator + TU'),
 ('','Isi TAB 4 (Jadwal Ngajugjug): tulis LOKASI NYATA tempat kerja tiap narasumber.','H-4','Koordinator'),
 ('','TEMUI SETIAP NARASUMBER secara langsung. Beri tahu: mereka akan didatangi 2 kelas, 10 menit masing-masing, pada Jumat 17 Juli, pada jam sekian. TEMPELKAN KARTU JADWAL kecil di tempat kerjanya. JANGAN LEWATKAN INI.','H-1 (Kamis 16 Juli)','Koordinator'),
 ('','Rekrut & briefing MARSHAL WALUYA: 12–16 murid OSIS/MPK. Sediakan rompi/pita lengan. Beri mereka daftar sektor 16 kelas.','H-2','Koordinator + Pembina OSIS'),
 ('','KOORDINASI DENGAN PANITIA MPLS KELAS X: sepakati slot wawancara adik kelas X (hanya saat istirahat MPLS), dan tegaskan ZONA MERAH.','H-3','Koordinator + Wakasek Kesiswaan'),
 ('','Briefing SATPAM: garis merah = tidak ada murid keluar gerbang.','H-1','Koordinator'),
 ('','Google Form "SUARA WALUYA": 1 tautan + 1 QR untuk 16 kelas. Dropdown "Kelas" berisi 16 pilihan. Akses Sheet diberikan ke Guru BK.','H-3','Tim ICT'),
 ('','Data sekunder dikumpulkan & dibagikan ke kelas yang membutuhkan: data sarpras/sanitasi (Wakasek Sarpras), data MBG (PJ MBG), data ekskul (Kesiswaan).','H-2','Koordinator'),
 ('','TEMPLATE DIGITAL TIM (02_UNTUK_MURID/TEMPLATE_DIGITAL_TIM.xlsx) diunggah ke Google Drive sekolah, lalu DISALIN menjadi 96 file (1 per tim) atau 16 file (1 per kelas). Bagikan tautannya ke tiap wali kelas. MURID TIDAK MENERIMA LEMBAR KERJA CETAK.','H-3','Tim ICT + Koordinator'),
 ('','Cetak PAPAN INSTRUKSI HARIAN (6 halaman) — 1 set per kelas, untuk ditayangkan atau ditempel guru.','H-2','Koordinator + TU'),
 ('','Cetak KARTU SAKU GURU (7 halaman A4) — 1 set per wali kelas. Sarankan dilaminating atau ditempel di meja.','H-2','Koordinator + TU'),
 ('','Ingatkan murid MEMBAWA BUKU TULIS khusus untuk JURNAL WALUYA (5 entri + Kontrak Kebiasaan). Tidak perlu beli baru.','H-1','Wali kelas'),
 ('','Alat tangan & bahan: 16 kotak (1 per kelas). Utamakan kardus & barang bekas. Anggaran Rp 200–400 rb/kelas.','H-2','Koordinator + Bendahara'),
 ('','Kotak P3K: 16 buah (1 per POS kelas).','H-2','UKS'),
 ('','BRIEFING 16 WALI KELAS (60 menit) — pakai 00_BRIEFING_GURU.pptx. Bagikan Kartu Sektor & jadwal gelombang masing-masing.','H-2','Koordinator'),
 ('','IZIN TITIK AKSI Hari 5: 16 titik pemasangan disetujui Wakasek Sarpras (satu per kelas, tidak boleh bentrok lokasi).','H-3','Koordinator + Wakasek Sarpras'),
 ('','Undang MITRA (Caraka, Satpam, Ibu Kantin, PJ MBG) ke Gelar Karya Hari 5. Mereka sudah dijanjikan murid — tepati.','H-1','Koordinator'),
 ('','Susun RUTE KUNJUNGAN Gelar Karya Lorong (searah, tidak saling tabrak). Tempel di tiap lorong.','H-2','Koordinator'),
]
r=5
for x in C:
    put(ws,r,list(x),h=38)
    if 'JANGAN LEWATKAN' in x[1] or 'ZONA MERAH' in x[1]:
        for i in range(1,5): ws.cell(row=r,column=i).fill=PatternFill('solid',fgColor=GOLD)
    r+=1
ws.freeze_panes='A5'
print('sheet7 ok')

wb.save(F); print('SAVED', F)
