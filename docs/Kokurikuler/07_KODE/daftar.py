import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
OUT="/sessions/vibrant-inspiring-einstein/mnt/Ko Kurikuler/SAKOLA_WALUYA_XI_2026/03_UNTUK_KOORDINATOR"
NAVY='1F3864'; HIJ='1FA971'; KUN='F0B429'; BIR='2E9BD6'; KOR='F4756B'; ORG='F2903D'
thin=Side(style='thin',color='B0BEC5'); BD=Border(left=thin,right=thin,top=thin,bottom=thin)

def th(ws,r,cols,widths,fill=NAVY,h=40):
    for i,(c,w) in enumerate(zip(cols,widths),1):
        x=ws.cell(r,i,c); x.font=Font(bold=True,color='FFFFFF',size=10)
        x.fill=PatternFill('solid',fgColor=fill); x.alignment=Alignment('center','center',wrap_text=True); x.border=BD
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[r].height=h
def title(ws,t,s,n):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n)
    c=ws.cell(1,1,t); c.font=Font(bold=True,size=15,color=NAVY); ws.row_dimensions[1].height=24
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=n)
    c=ws.cell(2,1,s); c.font=Font(italic=True,size=10,color='595959'); c.alignment=Alignment(wrap_text=True,vertical='top'); ws.row_dimensions[2].height=32
def put(ws,r,vals,h=110,fill=None):
    for i,v in enumerate(vals,1):
        x=ws.cell(r,i,v); x.border=BD; x.font=Font(size=10)
        x.alignment=Alignment(vertical='top',wrap_text=True)
        if i==1: x.alignment=Alignment('center','center')
        if fill and i==2: x.fill=PatternFill('solid',fgColor=fill); x.font=Font(size=10,bold=True)
    ws.row_dimensions[r].height=h

wb=openpyxl.Workbook()

# ══════ 1. DAFTAR KEGIATAN (siap input aplikasi) ══════
ws=wb.active; ws.title='1. DAFTAR KEGIATAN'
title(ws,'DAFTAR KEGIATAN KOKURIKULER — SMK NEGERI 2 CIMAHI · KELAS XI · TP 2026/2027',
 'Format mengikuti Panduan Kokurikuler 2025. Kolom "Profil Lulusan" diisi DIMENSI + SUB-DIMENSI (rumusan sub-dimensi diturunkan dari deskripsi 8 dimensi pada Panduan Kokurikuler 2025, hlm. 3–4).',6)
th(ws,4,['No','TEMA','JUDUL KEGIATAN KOKURIKULER','TUJUAN AKHIR KEGIATAN KOKURIKULER','PROFIL LULUSAN — DIMENSI','PROFIL LULUSAN — SUB-DIMENSI'],
   [5,26,26,54,24,52],NAVY,42)

DIM_UTAMA = 'kreativitas\nkolaborasi\nkesehatan'
SUB_UTAMA = ('kreativitas : Merumuskan solusi bagi permasalahan di sekitarnya; menciptakan inovasi\n\n'
             'kolaborasi : Peduli dan berbagi; membangun kerja sama dengan berbagai kalangan di lingkungan sekitar\n\n'
             'kesehatan : Hidup bersih dan sehat; berkontribusi secara positif terhadap lingkungan')

r=5
put(ws,r,[1,
  'Pendidikan Karakter Pancawaluya — Sekolah yang Cageur, Bageur, Bener, Pinter, tur Singer',
  'SAKOLA WALUYA — Ngamimitian ti Diri, Mere Mangpaat pikeun Sakola',
  'membiasakan tujuh kebiasaan anak Indonesia hebat serta menemukan, merancang, dan mewujudkan solusi nyata atas permasalahan lingkungan sekolah dengan menggunakan kompetensi keahlian yang dimiliki',
  DIM_UTAMA, SUB_UTAMA], h=190, fill='D6F2E4')

r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
c=ws.cell(r,1,'RINCIAN PER HARI  (opsional — dipakai bila aplikasi meminta kegiatan yang lebih granular)')
c.font=Font(bold=True,size=11,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='7F7F7F'); c.alignment=Alignment('left','center')
ws.row_dimensions[r].height=22

HARI=[
 (1,'CAGEUR — Kenali Diri','SAKOLA WALUYA Hari 1: Kenali Diri',
  'memetakan kondisi tujuh kebiasaan anak Indonesia hebat pada diri sendiri secara jujur serta menyepakati peran dan norma kerja tim',
  'kesehatan\nkolaborasi',
  'kesehatan : Hidup bersih dan sehat; kebugaran, kesehatan fisik, dan kesehatan mental\n\nkolaborasi : Peduli dan berbagi; membangun kerja sama dengan berbagai kalangan di lingkungan sekitar','D6F2E4'),
 (2,'BENER — Cari Fakta','SAKOLA WALUYA Hari 2: Cari Fakta',
  'mengumpulkan dan mengolah data lapangan di lingkungan sekolah serta merumuskan akar permasalahan yang didukung bukti',
  'penalaran kritis',
  'penalaran kritis : Rasa ingin tahu; berpikir logis dan analitis; menganalisis dan menyelesaikan permasalahan; berargumentasi logis; memanfaatkan literasi dan numerasi untuk memecahkan masalah','FFF3D6'),
 (3,'BAGEUR — Rasakan & Pilih','SAKOLA WALUYA Hari 3: Rasakan dan Pilih',
  'menggali perspektif warga sekolah melalui wawancara empatik dan menyepakati satu permasalahan prioritas secara demokratis',
  'kolaborasi\nkomunikasi',
  'kolaborasi : Peduli dan berbagi; membangun kerja sama dengan berbagai kalangan di lingkungan sekitar\n\nkomunikasi : Menyimak; berbicara dengan baik dan benar sesuai etika dalam beragam konteks dan moda','FDE6E4'),
 (4,'PINTER — Rancang & Buat','SAKOLA WALUYA Hari 4: Rancang dan Buat',
  'merancang dan mewujudkan purwarupa solusi atas permasalahan sekolah dengan menggunakan kompetensi keahlian program keahliannya',
  'kreativitas\nkolaborasi',
  'kreativitas : Berperilaku produktif; menciptakan inovasi; merumuskan solusi bagi permasalahan di sekitarnya\n\nkolaborasi : Peduli dan berbagi; membangun kerja sama dengan berbagai kalangan di lingkungan sekitar','DCEFFA'),
 (5,'SINGER — Bertindak & Berbagi','SAKOLA WALUYA Hari 5: Bertindak dan Berbagi',
  'mewujudkan solusi menjadi aksi nyata di lingkungan sekolah, mengomunikasikannya kepada warga sekolah, serta menyusun komitmen pembiasaan berkelanjutan',
  'kreativitas\nkomunikasi\nkemandirian',
  'kreativitas : Berperilaku produktif; menciptakan inovasi; merumuskan solusi bagi permasalahan di sekitarnya\n\nkomunikasi : Berbicara dan menulis dengan baik dan benar sesuai etika dalam beragam konteks dan moda\n\nkemandirian : Bertanggung jawab; berinisiatif; beradaptasi dalam pembelajaran dan pengembangan diri','FDEBDA'),
]
for n,tema,judul,tuj,dim,sub,col in HARI:
    r+=1
    put(ws,r,[n,tema,judul,tuj,dim,sub],h=120,fill=col)
ws.freeze_panes='A5'

# ══════ 2. REFERENSI 8 DIMENSI ══════
ws=wb.create_sheet('2. REFERENSI 8 DIMENSI')
title(ws,'REFERENSI — 8 DIMENSI PROFIL LULUSAN & SUB-DIMENSINYA',
 'Sumber: Panduan Kokurikuler 2025 (BSKAP Kemendikdasmen), hlm. 3–4. Sub-dimensi diturunkan langsung dari kalimat deskripsi tiap dimensi pada panduan tersebut.',4)
th(ws,4,['No','DIMENSI','DESKRIPSI DIMENSI (kutipan panduan)','SUB-DIMENSI (diturunkan dari deskripsi)'],[5,26,62,62],NAVY,36)
D8=[
 (1,'keimanan dan ketakwaan terhadap Tuhan Yang Maha Esa',
  'Mengacu pada individu yang memiliki keyakinan dan mengamalkan ajaran agama/kepercayaannya, berakhlak mulia, serta menjaga hubungan dengan Tuhan Yang Maha Esa, sesama manusia, dan lingkungan.',
  'Keyakinan dan pengamalan ajaran agama/kepercayaan\nAkhlak mulia\nHubungan dengan Tuhan Yang Maha Esa, sesama manusia, dan lingkungan'),
 (2,'kewargaan',
  'Mengacu pada individu yang bangga akan identitas dan budayanya, menghargai keberagaman, menjaga persatuan bangsa, menaati aturan bernegara dan bermasyarakat, serta menjaga keberlanjutan kehidupan, lingkungan, dan harmoni antarbangsa.',
  'Bangga akan identitas dan budaya\nMenghargai keberagaman\nMenjaga persatuan bangsa\nMenaati aturan bernegara dan bermasyarakat\nMenjaga keberlanjutan kehidupan, lingkungan, dan harmoni antarbangsa'),
 (3,'penalaran kritis',
  'Mengacu pada individu yang memiliki rasa ingin tahu, mampu berpikir logis dan analitis, serta mampu menganalisis dan menyelesaikan permasalahan, berargumentasi logis, dan memanfaatkan literasi dan numerasi untuk memecahkan masalah.',
  'Rasa ingin tahu\nBerpikir logis dan analitis\nMenganalisis dan menyelesaikan permasalahan\nBerargumentasi logis\nMemanfaatkan literasi dan numerasi untuk memecahkan masalah'),
 (4,'kreativitas',
  'Mengacu pada individu yang mampu berperilaku produktif, menciptakan inovasi, dan merumuskan solusi bagi permasalahan di sekitarnya.',
  'Berperilaku produktif\nMenciptakan inovasi\nMerumuskan solusi bagi permasalahan di sekitarnya'),
 (5,'kolaborasi',
  'Mengacu pada individu yang membiasakan diri untuk peduli dan berbagi, serta membangun kerja sama dengan berbagai kalangan di lingkungan sekitar.',
  'Peduli dan berbagi\nMembangun kerja sama dengan berbagai kalangan di lingkungan sekitar'),
 (6,'kemandirian',
  'Mengacu pada individu yang mampu bertanggung jawab, berinisiatif, dan beradaptasi dalam pembelajaran dan pengembangan diri.',
  'Bertanggung jawab\nBerinisiatif\nBeradaptasi dalam pembelajaran dan pengembangan diri'),
 (7,'kesehatan',
  'Mengacu pada individu yang menjalankan pola hidup bersih dan sehat berdasarkan pemahaman tentang kebugaran, kesehatan fisik dan mental, dan berkontribusi secara positif terhadap lingkungannya.',
  'Hidup bersih dan sehat\nKebugaran, kesehatan fisik, dan kesehatan mental\nBerkontribusi secara positif terhadap lingkungan'),
 (8,'komunikasi',
  'Mengacu pada individu yang memiliki kemampuan menyimak, membaca, berbicara, dan menulis dengan baik dan benar, sesuai etika dalam beragam konteks dan moda.',
  'Menyimak\nMembaca\nBerbicara\nMenulis dengan baik dan benar sesuai etika dalam beragam konteks dan moda'),
]
COLS=['E8F0FA','FFF3D6','D6F2E4','FDEBDA','DCEFFA','EEE7FA','FDE6E4','F7FBF9']
r=5
for (n,d,desc,sub),cl in zip(D8,COLS):
    put(ws,r,[n,d,desc,sub],h=(90 if len(sub)<120 else 110),fill=cl); r+=1
r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
c=ws.cell(r,1,'CATATAN: Panduan Kokurikuler 2025 TIDAK memuat daftar sub-dimensi baku. Rumusan sub-dimensi di atas diturunkan langsung dari kalimat deskripsi tiap dimensi pada panduan (hlm. 3–4) — cara yang sama dipakai aplikasi e-Rapor. Sesuaikan redaksinya dengan pilihan yang tersedia di aplikasi sekolah Anda.')
c.font=Font(size=10,italic=True,color='C00000'); c.alignment=Alignment(wrap_text=True,vertical='top'); ws.row_dimensions[r].height=50
ws.freeze_panes='A5'

# ══════ 3. CARA INPUT ══════
ws=wb.create_sheet('3. CARA INPUT')
ws.merge_cells('A1:B1'); c=ws.cell(1,1,'CARA MENGINPUT KE APLIKASI'); c.font=Font(bold=True,size=15,color=NAVY); ws.row_dimensions[1].height=26
ws.column_dimensions['A'].width=6; ws.column_dimensions['B'].width=115
LG=[
 'Buka menu Kegiatan Kokurikuler di aplikasi. Klik "Tambah".',
 'JUDUL KEGIATAN — salin dari kolom C lembar "1. DAFTAR KEGIATAN".',
 'TUJUAN AKHIR — salin dari kolom D. Ditulis dengan kata kerja, huruf kecil di awal (contoh aplikasi: "membina kesehatan jasmani dan rohani secara berkelanjutan...").',
 'PROFIL LULUSAN — centang DIMENSI pada kolom E, lalu pilih SUB-DIMENSI sesuai kolom F. Kalau daftar pilihan di aplikasi berbeda redaksinya, pilih yang paling dekat maknanya.',
 'Kalau aplikasi hanya menerima SATU kegiatan per semester — pakai BARIS 1 (SAKOLA WALUYA) saja. Kalau menerima kegiatan granular — pakai 5 baris rincian per hari.',
 'TEMA tidak selalu ada kolomnya di aplikasi. Kalau tidak ada, tema cukup disimpan di dokumen rancangan (kolom B lembar 1).',
]
for i,t in enumerate(LG,1):
    r=2+i
    x=ws.cell(r,1,i); x.font=Font(bold=True,size=13,color='FFFFFF'); x.fill=PatternFill('solid',fgColor=NAVY); x.alignment=Alignment('center','center'); x.border=BD
    y=ws.cell(r,2,t); y.font=Font(size=11); y.alignment=Alignment(vertical='center',wrap_text=True); y.border=BD
    ws.row_dimensions[r].height=46

wb.save(f'{OUT}/DAFTAR_KEGIATAN_KOKURIKULER.xlsx')
print('SAVED. sheets:',wb.sheetnames)
