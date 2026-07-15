import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
OUT="/sessions/vibrant-inspiring-einstein/mnt/Ko Kurikuler/SAKOLA_WALUYA_XI_2026/01_UNTUK_GURU"
NAVY='1F3864'; HIJ='1FA971'; KUN='F0B429'; BIR='2E9BD6'; KOR='F4756B'
thin=Side(style='thin',color='B0BEC5'); BD=Border(left=thin,right=thin,top=thin,bottom=thin)
IN=PatternFill('solid',fgColor='FFF9C4')

def th(ws,r,cols,widths,fill=NAVY):
    for i,(c,w) in enumerate(zip(cols,widths),1):
        x=ws.cell(r,i,c); x.font=Font(bold=True,color='FFFFFF',size=10)
        x.fill=PatternFill('solid',fgColor=fill); x.alignment=Alignment('center','center',wrap_text=True); x.border=BD
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[r].height=42

wb=openpyxl.Workbook()

# ══════════════ 0. KENAPA HANYA 3 DIMENSI ══════════════
ws=wb.active; ws.title='0. KENAPA 3 DIMENSI'
ws.merge_cells('A1:D1'); c=ws.cell(1,1,'KENAPA HANYA 3 DIMENSI YANG DINILAI?')
c.font=Font(bold=True,size=15,color=NAVY); ws.row_dimensions[1].height=26
ws.merge_cells('A2:D2')
c=ws.cell(2,1,'Panduan Kokurikuler 2025 tidak mewajibkan semua dimensi dinilai. Contoh rubrik di panduan justru hanya memakai 2 dimensi. '
              'Dimensi dipilih dari titik TERLEMAH pada Rapor Pendidikan SMKN 2 Cimahi — bukan dari selera. '
              'Lima dimensi lainnya DISIMPAN untuk projek kokurikuler berikutnya, supaya tidak diborong sekaligus dan penilaian tetap jujur.')
c.font=Font(italic=True,size=10,color='595959'); c.alignment=Alignment(wrap_text=True,vertical='top'); ws.row_dimensions[2].height=46

th(ws,4,['DIMENSI','BUKTI DI RAPOR PENDIDIKAN','KEPUTUSAN','ALASAN'],[20,42,20,58])
PILIH=[
 ('kreativitas','A.3.3 Kreativitas = 54,31 — TERENDAH dari 6 sub-dimensi karakter.','DINILAI',
  'Titik paling lemah. Seluruh Hari 4-5 memang dirancang untuk ini: merancang & mewujudkan solusi.', HIJ),
 ('kolaborasi','A.3.2 Gotong Royong = 54,70 — SATU-SATUNYA sub-dimensi karakter yang TURUN (-3,09).','DINILAI',
  'Satu-satunya yang memburuk. Struktur 6 peran murid & Lembar Silih Asah langsung menyasar ini.', BIR),
 ('kesehatan','D.19 Tujuh Kebiasaan = 6,57 — 5 dari 7 kebiasaan berstatus KURANG.','DINILAI',
  'Tulang punggung kegiatan (7 KAIH). Diukur lewat Radar Kebiasaan Hari 1 vs Hari 5 — ada bukti angka.', KOR),
 ('penalaran kritis','A.3.4 = 56,75 — naik 0,23 (nyaris stagnan, tapi BUKAN yang terburuk).','tidak dinilai formal',
  'Tetap DILATIH intensif di Hari 2 (data, pohon akar). Tapi tidak dimasukkan rapor supaya guru tidak kewalahan. Simpan untuk projek berikutnya.', 'BFBFBF'),
 ('komunikasi','—','tidak dinilai formal','Tersentuh lewat presentasi 60 detik & wawancara narasumber. Simpan untuk projek berikutnya.', 'BFBFBF'),
 ('kemandirian','A.3.6 = 58,50 — naik 3,88 (sudah membaik sendiri).','tidak dinilai formal',
  'Tersentuh lewat Kontrak Kebiasaan. Tidak mendesak.', 'BFBFBF'),
 ('kewargaan','—','tidak dinilai formal','Tersentuh lewat penyerahan resmi gambar kerja ke Sarpras. Simpan untuk projek berikutnya.', 'BFBFBF'),
 ('keimanan dan ketakwaan','A.3.1 = 59,72 — TERTINGGI.','tidak dinilai formal',
  'Sudah paling baik. Tersentuh lewat apel & Jumat Berkah.', 'BFBFBF'),
]
r=5
for d,b,k,a,col in PILIH:
    for i,v in enumerate([d,b,k,a],1):
        x=ws.cell(r,i,v); x.border=BD; x.alignment=Alignment(vertical='center',wrap_text=True); x.font=Font(size=10)
        if i==1:
            x.font=Font(size=11,bold=True,color='FFFFFF'); x.fill=PatternFill('solid',fgColor=col)
            x.alignment=Alignment('center','center',wrap_text=True)
        if i==3:
            x.font=Font(size=10,bold=True, color=('1F3864' if k=='DINILAI' else '7F7F7F'))
            x.alignment=Alignment('center','center',wrap_text=True)
    ws.row_dimensions[r].height=46; r+=1
r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
c=ws.cell(r,1,'AKIBATNYA UNTUK WALI KELAS: hanya 3 kolom nilai per murid (bukan 8). Sub-dimensi TIDAK dinilai terpisah — hanya jadi panduan mengamati.')
c.font=Font(bold=True,size=11,color=NAVY); c.fill=PatternFill('solid',fgColor='FFF3D6'); c.border=BD
c.alignment=Alignment(wrap_text=True,vertical='center'); ws.row_dimensions[r].height=32

# ══════════════ 1. RUBRIK ══════════════
ws=wb.create_sheet('1. RUBRIK')
ws.merge_cells('A1:G1'); c=ws.cell(1,1,'RUBRIK PENILAIAN KINERJA — SAKOLA WALUYA')
c.font=Font(bold=True,size=15,color=NAVY); ws.row_dimensions[1].height=24
ws.merge_cells('A2:G2'); c=ws.cell(2,1,'Format mengikuti Panduan Kokurikuler 2025: Dimensi Profil Lulusan x Aspek yang Dinilai x SB/B/C/K. Dipakai pada GELAR KARYA, Hari 5. '
     'Kolom SUB-DIMENSI hanya PANDUAN MENGAMATI — beri SATU nilai per DIMENSI, bukan per sub-dimensi.')
c.font=Font(italic=True,size=10,color='595959'); c.alignment=Alignment(wrap_text=True,vertical='top'); ws.row_dimensions[2].height=32
th(ws,4,['Dimensi\nProfil Lulusan','Sub-dimensi yang diamati\n(rujukan: Panduan Kokurikuler)','Aspek yang Dinilai','SANGAT BAIK (SB)','BAIK (B)','CUKUP (C)','KURANG (K)'],[16,30,28,32,28,28,28])
RUB=[
 ('kreativitas',
  '1. Merumuskan solusi bagi permasalahan di sekitarnya\n\n2. Menciptakan inovasi',
  'Merancang dan mewujudkan solusi atas masalah nyata sekolah dengan kompetensi keahliannya',
  'Solusi orisinal, memanfaatkan kompetensi keahlian secara maksimal, dan benar-benar berfungsi/terpasang di sekolah.',
  'Solusi relevan dan memakai kompetensi keahlian; purwarupa berfungsi meski sederhana.',
  'Solusi meniru contoh yang diberikan; kaitannya dengan kompetensi keahlian masih tipis.',
  'Tidak menghasilkan solusi/karya, atau karyanya tidak berhubungan dengan masalah yang dipilih.', HIJ),
 ('kolaborasi',
  '1. Peduli dan berbagi\n\n2. Membangun kerja sama dengan berbagai kalangan di lingkungan sekitar',
  'Menjalankan peran dalam tim, dan membantu tim lain',
  'Menjalankan perannya secara konsisten tanpa diingatkan, menyelesaikan konflik dalam timnya sendiri, dan membantu tim lain (silih asih-asah-asuh).',
  'Menjalankan perannya dengan baik dan berkontribusi nyata pada hasil kerja tim.',
  'Menjalankan perannya bila diingatkan; kontribusinya belum merata.',
  'Tidak menjalankan perannya; pekerjaan tim ditanggung anggota lain.', BIR),
 ('kesehatan',
  '1. Hidup bersih dan sehat\n\n2. Berkontribusi secara positif terhadap lingkungan',
  'Menjalankan pembiasaan 7 KAIH dan berkontribusi pada kesehatan/kebersihan sekolah',
  'Menjalankan pembiasaan 7 KAIH selama lima hari (jurnal terisi penuh), skor Radar Kebiasaan berubah, dan karyanya berdampak nyata pada kesehatan/kebersihan sekolah.',
  'Menjalankan sebagian besar pembiasaan; jurnal terisi; karyanya berkaitan dengan kesehatan/kebersihan sekolah.',
  'Menjalankan pembiasaan bila diingatkan; jurnal terisi sebagian.',
  'Tidak menjalankan pembiasaan; jurnal kosong atau diisi asal-asalan.', KOR),
]
r=5
for d,sd,a,sb,b,cc,k,col in RUB:
    for i,v in enumerate([d,sd,a,sb,b,cc,k],1):
        x=ws.cell(r,i,v); x.border=BD; x.alignment=Alignment(vertical='top',wrap_text=True); x.font=Font(size=10)
        if i==1:
            x.font=Font(size=11,bold=True,color='FFFFFF'); x.fill=PatternFill('solid',fgColor=col)
            x.alignment=Alignment('center','center',wrap_text=True)
        if i==2: x.font=Font(size=10); x.fill=PatternFill('solid',fgColor='F2F5F6')
        if i==3: x.font=Font(size=10,bold=True)
    ws.row_dimensions[r].height=110; r+=1
r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
c=ws.cell(r,1,'Sub-dimensi di atas dikutip dari deskripsi dimensi pada Panduan Kokurikuler 2025 (hlm. 3-4). Panduan tidak memuat daftar sub-dimensi baku; '
              'kalau daftar di aplikasi rapor sekolah berbeda redaksinya, pilih yang paling dekat maknanya.')
c.font=Font(italic=True,size=10,color='595959'); c.alignment=Alignment(wrap_text=True); ws.row_dimensions[r].height=30
r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
c=ws.cell(r,1,'Dimensi penyerta (DILATIH, tetapi TIDAK diasesmen formal): penalaran kritis (Hari 2) - komunikasi (presentasi 60 detik) - '
              'kemandirian (Kontrak Kebiasaan) - kewargaan (penyerahan resmi gambar kerja) - keimanan & ketakwaan (apel, Jumat Berkah). '
              'Semuanya disimpan untuk projek kokurikuler berikutnya.')
c.font=Font(italic=True,size=10,color='595959'); c.alignment=Alignment(wrap_text=True); ws.row_dimensions[r].height=30

# ══════════════ 2. LEMBAR NILAI ══════════════
ws=wb.create_sheet('2. LEMBAR NILAI')
ws.merge_cells('A1:G1'); c=ws.cell(1,1,'LEMBAR PENILAIAN — KELAS: ......................')
c.font=Font(bold=True,size=15,color=NAVY); ws.row_dimensions[1].height=24
ws.merge_cells('A2:G2'); c=ws.cell(2,1,'Isi kolom C, D, E dengan SB / B / C / K (pilih dari daftar). HANYA 3 KOLOM NILAI. Kolom G menyusun DRAF DESKRIPSI RAPOR otomatis — tinggal dirapikan.')
c.font=Font(italic=True,size=10,color='595959'); ws.row_dimensions[2].height=18
th(ws,4,['No','NAMA MURID','kreativitas','kolaborasi','kesehatan','Catatan Pendidik\n(opsional)','DRAF DESKRIPSI RAPOR (otomatis)'],[5,26,13,13,13,36,80])

dv=DataValidation(type="list", formula1='"SB,B,C,K"', allow_blank=True)
ws.add_data_validation(dv)
MAP = 'IF({c}="SB","sangat baik",IF({c}="B","baik",IF({c}="C","cukup",IF({c}="K","perlu bimbingan",""))))'
for i in range(36):
    r = 5 + i
    ws.cell(r,1,i+1).border = BD; ws.cell(r,1).alignment = Alignment('center')
    for cx in range(2,7):
        x=ws.cell(r,cx); x.border=BD; x.fill=IN; x.alignment=Alignment(vertical='center',wrap_text=True); x.font=Font(size=10)
    for cx in range(3,6):
        dv.add(ws.cell(r,cx)); ws.cell(r,cx).alignment=Alignment('center','center')
    f = ('=IF(B{r}="","","Ananda "&B{r}&" menunjukkan capaian "'
         '&{mc}&" dalam kreativitas, "&{md}&" dalam kolaborasi, dan "&{me}'
         '&" dalam kesehatan, pada kegiatan kokurikuler Sakola Waluya."'
         '&IF(F{r}=""," "," "&F{r}))').format(
            r=r, mc=MAP.format(c='C%d'%r), md=MAP.format(c='D%d'%r), me=MAP.format(c='E%d'%r))
    x=ws.cell(r,7,f); x.border=BD; x.alignment=Alignment(vertical='center',wrap_text=True); x.font=Font(size=10)
    ws.row_dimensions[r].height=40
ws.freeze_panes='A5'

# ══════════════ 3. CATATAN ANEKDOTAL ══════════════
ws=wb.create_sheet('3. CATATAN ANEKDOTAL')
ws.merge_cells('A1:E1'); c=ws.cell(1,1,'CATATAN ANEKDOTAL (asesmen formatif) — KELAS: ......................')
c.font=Font(bold=True,size=15,color=NAVY); ws.row_dimensions[1].height=24
ws.merge_cells('A2:E2'); c=ws.cell(2,1,'Target realistis 3 catatan per hari (guru sendirian). Catat FAKTA yang terlihat/terdengar — BUKAN tafsir. Salah: "Rizky malas." Benar: "Rizky tidak mengambil peran selama 30 menit dan menatap HP."')
c.font=Font(italic=True,size=10,color='595959'); c.alignment=Alignment(wrap_text=True,vertical='top'); ws.row_dimensions[2].height=30
th(ws,4,['Tanggal','Nama Murid','Perilaku yang Teramati (fakta, bukan penilaian)','Dimensi','Nilai Pancawaluya'],[12,22,60,20,20])
dv2=DataValidation(type="list", formula1='"kreativitas,kolaborasi,kesehatan"', allow_blank=True)
ws.add_data_validation(dv2)
ex=['16/07','Contoh: Rizky','Saat timnya kekurangan data, ia meminjamkan meteran ke tim lain lebih dulu sebelum timnya selesai mengukur.','kolaborasi','BAGEUR — silih asuh']
for i,v in enumerate(ex,1):
    x=ws.cell(5,i,v); x.border=BD; x.font=Font(size=10,italic=True,color='7F7F7F'); x.alignment=Alignment(vertical='center',wrap_text=True)
ws.row_dimensions[5].height=48
for i in range(30):
    r=6+i
    for cx in range(1,6):
        x=ws.cell(r,cx); x.border=BD; x.fill=IN; x.font=Font(size=10); x.alignment=Alignment(vertical='center',wrap_text=True)
    dv2.add(ws.cell(r,4)); ws.row_dimensions[r].height=32
ws.freeze_panes='A5'

# ══════════════ 4. CARA MENGISI ══════════════
ws=wb.create_sheet('4. CARA MENGISI')
ws.merge_cells('A1:B1'); c=ws.cell(1,1,'CARA MENGISI — 5 LANGKAH'); c.font=Font(bold=True,size=15,color=NAVY); ws.row_dimensions[1].height=26
ws.column_dimensions['A'].width=6; ws.column_dimensions['B'].width=110
LG=[
 'SEPANJANG KEGIATAN — isi lembar "3. CATATAN ANEKDOTAL". Cukup 3 catatan per hari. Ini asesmen FORMATIF: untuk umpan balik, bukan untuk menghakimi.',
 'SETIAP MALAM — baca Jurnal Waluya murid (buku tulis), balas SATU kalimat di tiap buku. Bukan nilai, bukan koreksi.',
 'HARI 5, SAAT GELAR KARYA — nilai tiap murid dengan "1. RUBRIK". HANYA 3 DIMENSI: kreativitas, kolaborasi, kesehatan. Lingkari SB/B/C/K. Penilai: Anda (utama) + Kaprog (penilai teknis tamu) + OSIS.',
 'MASUKKAN HASILNYA ke lembar "2. LEMBAR NILAI" (kolom C, D, E). Kolom G otomatis menyusun DRAF deskripsi rapor.',
 'RAPIKAN draf itu. Panduan Kokurikuler 2025: ringkas, POSITIF, edukatif. Sebut yang sudah baik LEBIH DULU, lalu yang masih perlu dilatih. Hindari kata "kurang" atau "gagal". Salin ke kolom KOKURIKULER di rapor.',
]
for i,t in enumerate(LG,1):
    r=2+i
    x=ws.cell(r,1,i); x.font=Font(bold=True,size=13,color='FFFFFF'); x.fill=PatternFill('solid',fgColor=NAVY); x.alignment=Alignment('center','center'); x.border=BD
    y=ws.cell(r,2,t); y.font=Font(size=11); y.alignment=Alignment(vertical='center',wrap_text=True); y.border=BD
    ws.row_dimensions[r].height=48
r=9
ws.cell(r,2,'CONTOH DESKRIPSI RAPOR YANG SUDAH DIRAPIKAN (3 dimensi)').font=Font(bold=True,size=12,color=NAVY); r+=1
for t in ['"Ananda Rizky sudah sangat baik dalam KOLABORASI dengan menjalankan perannya secara konsisten dan membantu tim lain, serta baik dalam KESEHATAN melalui pembiasaan tujuh kebiasaan yang terjaga selama kegiatan. Masih perlu berlatih dalam KREATIVITAS, khususnya mengembangkan gagasan sendiri, tidak sekadar meniru contoh."',
          '"Ananda Sinta sudah sangat baik dalam KREATIVITAS saat merancang dispenser sabun mekanis yang kini terpasang di sekolah, serta baik dalam KOLABORASI dan KESEHATAN."']:
    x=ws.cell(r,2,t); x.font=Font(size=11,italic=True); x.alignment=Alignment(vertical='center',wrap_text=True)
    x.fill=PatternFill('solid',fgColor='FFF3D6'); x.border=BD; ws.row_dimensions[r].height=62; r+=1
r+=1
x=ws.cell(r,2,'CATATAN: kalau ada murid yang menonjol di penalaran kritis atau komunikasi, tulis saja di kolom "Catatan Pendidik" — akan ikut masuk ke draf deskripsi. Tidak perlu kolom nilai tersendiri.')
x.font=Font(size=11,bold=True,color=NAVY); x.alignment=Alignment(vertical='center',wrap_text=True); ws.row_dimensions[r].height=40

wb.save(f'{OUT}/INSTRUMEN_ASESMEN.xlsx')
print('SAVED. sheets:', wb.sheetnames)
