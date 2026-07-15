import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = "/sessions/vibrant-inspiring-einstein/mnt/Ko Kurikuler/SAKOLA_WALUYA_XI_2026/03_UNTUK_KOORDINATOR"
os.makedirs(OUT, exist_ok=True)

NAVY='1F3864'; BLUE='2E74B5'; BIRU='DEEBF7'; JINGGA='FCE4D6'
Z={'Z1':'DDEBF7','Z2':'FFF2CC','Z3':'E4DFEC','Z4':'E2EFDA','Z5':'FCE4EC'}
thin=Side(style='thin',color='B0BEC5')
BD=Border(left=thin,right=thin,top=thin,bottom=thin)

def hdr(ws,row,cols,fill=NAVY):
    for i,c in enumerate(cols,1):
        cell=ws.cell(row=row,column=i,value=c)
        cell.font=Font(bold=True,color='FFFFFF',size=10,name='Calibri')
        cell.fill=PatternFill('solid',fgColor=fill)
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        cell.border=BD
    ws.row_dimensions[row].height=32

def title(ws,text,sub,ncol):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol)
    c=ws.cell(row=1,column=1,value=text)
    c.font=Font(bold=True,size=15,color=NAVY,name='Calibri'); c.alignment=Alignment(horizontal='left',vertical='center')
    ws.row_dimensions[1].height=24
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncol)
    c=ws.cell(row=2,column=1,value=sub)
    c.font=Font(italic=True,size=10,color='595959',name='Calibri')
    ws.row_dimensions[2].height=18

def w(ws,widths):
    for i,x in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=x

def put(ws,r,vals,fill=None,bold=False,wrap=True,size=10):
    for i,v in enumerate(vals,1):
        c=ws.cell(row=r,column=i,value=v)
        c.font=Font(size=size,bold=bold,name='Calibri')
        c.alignment=Alignment(vertical='center',wrap_text=wrap,horizontal='left')
        c.border=BD
        if fill: c.fill=PatternFill('solid',fgColor=fill)

wb=openpyxl.Workbook()

# ============ 1. MASTER ============
ws=wb.active; ws.title='1. MASTER 16 KELAS'
title(ws,'MATRIKS OPERASIONAL — SAKOLA WALUYA · 16 KELAS PARALEL','SMK Negeri 2 Cimahi · Kelas XI · 15–21 Juli 2026 · Setiap kelas punya SEKTOR EKSKLUSIF & GELOMBANG. Tidak ada dua kelas yang mengukur objek yang sama pada waktu yang sama.',8)
hdr(ws,4,['No','KELAS','GELOMBANG','ZONA AKSI','SEKTOR EKSKLUSIF\n(objek data yang HANYA boleh diukur kelas ini)','NARASUMBER KUNCI\n(didatangi delegasi, 10 menit)','BENTUK KARYA (tanpa bengkel)','LORONG GELAR KARYA'])
w(ws,[4,24,12,22,44,30,46,20])

DATA=[
 (1,'XI MEKATRONIKA - A','BIRU','Z1 Cai & Kabersihan','Toilet & wastafel GEDUNG/BLOK A','N1 · Petugas Kebersihan (Caraka) 1','Dispenser sabun tekan/pedal dari botol bekas (mekanis, tanpa listrik) + gambar teknis','Depan kelas'),
 (2,'XI KIMIA INDUSTRI - A','BIRU','Z1 Cai & Kabersihan','Toilet & wastafel GEDUNG/BLOK B','N2 · Petugas Kebersihan (Caraka) 2','Uji kualitas air: indikator alami (kubis ungu/kunyit) + pH strip + uji kekeruhan & bau','Depan kelas'),
 (3,'XI KIMIA INDUSTRI - B','JINGGA','Z1 Cai & Kabersihan','Toilet & wastafel GEDUNG/BLOK C + sumber air / tandon','N1 · Petugas Kebersihan (Caraka) 1','Uji air sumber/tandon + pembuatan cairan pembersih ramah lingkungan (eco-enzyme)','Depan kelas'),
 (4,'XI TEKNIK PEMESINAN - A','JINGGA','Z1 Cai & Kabersihan','Toilet & wastafel GEDUNG/BLOK D + saluran air','N2 · Petugas Kebersihan (Caraka) 2','Purwarupa 1:1 kardus/PVC: dudukan & rak wastafel + GAMBAR KERJA untuk difabrikasi','Depan kelas'),
 (5,'XI DKV - A','BIRU','Z2 Dahar Sehat','KANTIN — jajanan, harga, kandungan gula, antrean','N3 · Ibu/Bapak Kantin','Infografis gizi jajanan + signage kantin + ikon pemilahan sampah','Depan kelas'),
 (6,'XI KIMIA INDUSTRI - C','BIRU','Z2 Dahar Sehat','SAMPAH MAKANAN — timbang sisa MBG seluruh kelas XI + tempat sampah sekolah','N4 · Penanggung Jawab MBG / petugas SPPG','ECO-ENZYME dari sisa kulit buah & sayur MBG → cairan pembersih + audit food waste 5 hari','Depan kelas'),
 (7,'XI RPL - A','JINGGA','Z2 Dahar Sehat','DISTRIBUSI MBG — alur pengambilan, waktu antre, keterserapan porsi','N3 · Ibu/Bapak Kantin','Sistem catat & rekap MBG (Google Form + Sheet + dasbor keterserapan porsi)','Depan kelas'),
 (8,'XI MEKATRONIKA - D','BIRU','Z3 Betah Diajar','RUANG KELAS — pencahayaan, ventilasi, kebisingan, kenyamanan belajar','N5 · Petugas TU / Perpustakaan','Purwarupa lampu belajar / pengingat waktu tidur (baterai) atau simulasi Tinkercad + gambar teknis','Depan kelas'),
 (9,'XI DKV - B','JINGGA','Z3 Betah Diajar','PERPUSTAKAAN & POJOK BACA — ketersediaan, kondisi, pemanfaatan','N4 · Penanggung Jawab MBG / petugas SPPG','Signage & wayfinding pojok baca + poster kampanye Gemar Belajar','Depan kelas'),
 (10,'XI RPL - B','JINGGA','Z3 Betah Diajar','RITME TIDUR & GAWAI — survei daring + observasi kantuk di kelas (jam ke-1 vs ke-7)','N5 · Petugas TU / Perpustakaan','Tracker kebiasaan 7 KAIH berbasis web/Sheet + dasbor kelas','Depan kelas'),
 (11,'XI MEKATRONIKA - B','BIRU','Z4 Awak Bugar','LAPANGAN & SARANA OLAHRAGA — kondisi, pemanfaatan, jam pakai','N8 · Tukang Kebun / Penjaga Parkir','Purwarupa timer/bel "Jeda Gerak" (mekanis atau baterai) + gambar teknis','Depan kelas'),
 (12,'XI ANIMASI - A','BIRU','Z4 Awak Bugar','AKTIVITAS FISIK HARIAN — hitung langkah & durasi duduk (sampel murid, pakai HP)','N7 · Pembina Ekskul / Kesiswaan','Video/animasi "Jeda Gerak 3 Menit" untuk ditayangkan di kelas + PSA kebugaran','Depan kelas'),
 (13,'XI MEKATRONIKA - C','JINGGA','Z4 Awak Bugar','KORIDOR & TANGGA — jalur gerak, hambatan, potensi titik "jeda gerak"','N7 · Pembina Ekskul / Kesiswaan','Purwarupa penanda/alat bantu aktivitas fisik di koridor (mekanis) + gambar teknis','Depan kelas'),
 (14,'XI ANIMASI - B','BIRU','Z5 Sakola Aman','PETA TITIK RAWAN LANTAI 1 & AREA LUAR (parkir, belakang) — foto LOKASI saja, BUKAN orang','N6 · Satpam','PSA animasi 30–60 detik anti-perundungan + video "Kalau kamu tidak aman, lapor ke mana"','Depan kelas'),
 (15,'XI DKV - C','JINGGA','Z5 Sakola Aman','PETA TITIK RAWAN LANTAI 2 & ATAS — foto LOKASI saja, BUKAN orang','N6 · Satpam','Papan informasi KANAL PELAPORAN + kampanye visual sekolah aman','Depan kelas'),
 (16,'XI TEKNIK PEMESINAN - B','JINGGA','Z5 Sakola Aman','PAPAN INFORMASI & KANAL PENGADUAN — inventarisasi seluruh sekolah (adakah? di mana? berfungsi?)','N8 · Tukang Kebun / Penjaga Parkir','Purwarupa 1:1 kardus/PVC: KOTAK PENGADUAN + papan informasi + GAMBAR KERJA untuk difabrikasi','Depan kelas'),
]
r=5
for d in DATA:
    fill = BIRU if d[2]=='BIRU' else JINGGA
    put(ws,r,list(d),fill=fill); ws.row_dimensions[r].height=46; r+=1

r+=1
ws.cell(row=r,column=1,value='CATATAN PENTING').font=Font(bold=True,size=11,color='C00000',name='Calibri')
notes=[
 '1. "GEDUNG/BLOK A–D" adalah PENANDA SEMENTARA. Koordinator WAJIB memetakannya ke denah nyata SMKN 2 Cimahi paling lambat H-3, lalu menuliskannya di Kartu Sektor tiap kelas.',
 '2. GELOMBANG BIRU turun lapangan LEBIH DULU; GELOMBANG JINGGA bekerja di kelas, lalu bertukar. Maksimal 8 kelas (48 tim) di lapangan pada satu waktu — bukan 16 kelas (96 tim).',
 '3. SEKTOR bersifat EKSKLUSIF. Tim dilarang keras mengukur/menjelajah objek di sektor kelas lain. Kalau butuh data dari sektor lain — MINTA ke kelas itu (ini gotong royong nyata).',
 '4. NARASUMBER TIDAK DIKUMPULKAN. Mereka tetap bekerja seperti biasa di tempatnya. MURID yang MENGHAMPIRI (ngajugjug) — tapi HANYA delegasi 3 orang per kelas, HANYA 10 menit, HANYA pada slot yang dijadwalkan. Lihat tab "4. JADWAL NGAJUGJUG".',
 '5. ZONA MERAH: area MPLS kelas X. Tim XI DILARANG masuk kecuali pada slot wawancara adik kelas X yang sudah dikoordinasikan dengan panitia MPLS. Lihat tab "6. MARSHAL & ZONA MERAH".',
 '6. Kelas XII sedang PKL / belum masuk — ruang & koridor kelas XII boleh dipakai sebagai area cadangan.',
]
for n in notes:
    r+=1
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
    c=ws.cell(row=r,column=1,value=n); c.font=Font(size=10,name='Calibri'); c.alignment=Alignment(wrap_text=True,vertical='top')
    ws.row_dimensions[r].height=30
ws.freeze_panes='A5'
print('sheet1 ok')

# ============ 2. JADWAL HARI 2 ============
ws=wb.create_sheet('2. JADWAL HARI 2')
title(ws,'JADWAL HARI 2 — KAMIS, 16 JULI 2026 · BENER · "CARI FAKTA"','Sistem GELOMBANG. Maksimal 8 kelas (48 tim) di lapangan pada satu waktu.',4)
hdr(ws,4,['WAKTU','GELOMBANG BIRU (8 kelas)','GELOMBANG JINGGA (8 kelas)','KETERANGAN'])
w(ws,[18,50,50,38])
J2=[
 ('06.30–06.45','APEL PENGONDISIAN (lapangan, seluruh kelas XI)','APEL PENGONDISIAN','Bangun Pagi · Beribadah'),
 ('06.45–07.30','MBG BERSAMA + NGARIUNG (di kelas)','MBG BERSAMA + NGARIUNG','Makan Sehat · Bermasyarakat. Tim piket MENIMBANG sisa makanan kelas → tulis di papan (data untuk XI KIMIA C)'),
 ('07.30–07.45','PEMANTIK "Katanya vs Buktinya" (di kelas)','PEMANTIK "Katanya vs Buktinya"','Menanamkan nilai BENER'),
 ('07.45–09.15\n(90 menit)','TURUN LAPANGAN — SEKTOR MASING-MASING\n• Guru menetap di POS\n• Tim Piket Data menjaga PAPAN KARTU\n• Lapor balik tiap 20 menit','DI KELAS — PERSIAPAN & DATA SEKUNDER\n• Pelajari denah sektor & tandai objek\n• Susun instrumen ukur\n• Olah DATA SEKUNDER dari Wakasek Sarpras & PJ MBG\n• Susun HIPOTESIS yang akan diuji','Hanya 8 kelas (48 tim) di lapangan'),
 ('09.15–09.30','JEDA GERAK & SNACK (wajib berdiri)','JEDA GERAK & SNACK','Berolahraga'),
 ('09.30–11.00\n(90 menit)','DI KELAS — OLAH DATA\n• Hitung total, rata-rata, persentase\n• GAMBAR TANGAN grafik di plano\n• Galeri Data antartim','TURUN LAPANGAN — SEKTOR MASING-MASING\n• Guru menetap di POS\n• Tim Piket Data menjaga PAPAN KARTU\n• Lapor balik tiap 20 menit','Bertukar. Hanya 8 kelas di lapangan'),
 ('11.00–11.45\n(45 menit)','POHON AKAR MASALAH (5x KENAPA) + presentasi kilat antartim + DAFTAR MASALAH KELAS','OLAH DATA → grafik + POHON AKAR MASALAH (dipercepat: hipotesis & data sekunder sudah siap sejak pagi)','—'),
 ('11.45–12.00','REFLEKSI & JURNAL WALUYA','REFLEKSI & JURNAL WALUYA','Gemar Belajar (refleksi)'),
]
r=5
for x in J2:
    put(ws,r,list(x)); ws.row_dimensions[r].height=(78 if '•' in x[1] or '•' in x[2] else 34)
    ws.cell(row=r,column=2).fill=PatternFill('solid',fgColor=BIRU)
    ws.cell(row=r,column=3).fill=PatternFill('solid',fgColor=JINGGA)
    r+=1
ws.freeze_panes='A5'
print('sheet2 ok')

# ============ 3. JADWAL HARI 3 ============
ws=wb.create_sheet('3. JADWAL HARI 3')
title(ws,'JADWAL HARI 3 — JUMAT, 17 JULI 2026 · BAGEUR · "RASAKAN & PILIH"  (BERAKHIR 11.00)','Narasumber TIDAK dikumpulkan. Murid yang MENGHAMPIRI (ngajugjug) — hanya delegasi 3 orang, 10 menit, sesuai slot.',4)
hdr(ws,4,['WAKTU','GELOMBANG BIRU (8 kelas)','GELOMBANG JINGGA (8 kelas)','KETERANGAN'])
w(ws,[18,50,50,38])
J3=[
 ('06.30–07.30','APEL + MBG BERSAMA + NGARIUNG','APEL + MBG BERSAMA + NGARIUNG','Pertanyaan Ngariung: "Siapa orang di sekolah ini yang kerjanya paling tidak kelihatan, tapi paling kita butuhkan?"'),
 ('07.30–07.45','PEMANTIK "Siapa yang Paling Merasakan?"','PEMANTIK + finalisasi Pohon Akar Masalah (kalau belum selesai)','Memindahkan fokus dari MASALAH ke MANUSIA'),
 ('07.45–08.00\n(15 menit)','LATIHAN BERTANYA + bagi KARTU TEMU NARASUMBER + briefing ETIKA','LATIHAN BERTANYA + bagi KARTU TEMU NARASUMBER + briefing ETIKA','Semua kelas. Pertanyaan BURUK vs BAIK. Aturan: TANYA → DENGAR → JANGAN MEMOTONG → TANYA LAGI "kenapa"'),
 ('08.00–09.00\n(60 menit)','TURUN LAPANGAN\n• DELEGASI (3 murid) NGAJUGJUG narasumber kunci di TEMPAT KERJANYA — 10 menit saja, sesuai slot\n• Tim lain mewawancarai warga sekolah biasa di SEKTOR: murid kelas X (MPLS), murid kelas lain, guru piket','DI KELAS\n• Susun daftar pertanyaan wawancara\n• Peta Empati awal (dari data Hari 2)\n• Selesaikan olahan data yang tertunda','Hanya 8 kelas di lapangan. Slot narasumber BIRU: lihat tab 4'),
 ('09.00–09.15','JEDA GERAK & SNACK','JEDA GERAK & SNACK','—'),
 ('09.15–10.15\n(60 menit)','DI KELAS\n• PETA EMPATI (4 kuadran)\n• MATRIKS PRIORITAS\n• VOTING masalah prioritas (voting kaki, 3 stiker/murid)','TURUN LAPANGAN\n• DELEGASI (3 murid) NGAJUGJUG narasumber kunci\n• Tim lain wawancara warga sekolah biasa di SEKTOR','Bertukar. Slot narasumber JINGGA: lihat tab 4'),
 ('10.15–10.45\n(30 menit)','FINALISASI: tulis plano besar "MASALAH KAMI / ORANG YANG KAMI BANTU / TARGET SELESAI SELASA" + siapkan Jumat Berkah','PETA EMPATI + MATRIKS PRIORITAS + VOTING + plano besar','Keputusan diambil MURID'),
 ('10.45–11.00','JUMAT BERKAH + REFLEKSI & JURNAL','JUMAT BERKAH + REFLEKSI & JURNAL','Murid MENYERAHKAN temuan + ucapan terima kasih tertulis kepada narasumber. Lalu dilepas ke Shalat Jumat.'),
]
r=5
for x in J3:
    put(ws,r,list(x)); ws.row_dimensions[r].height=(84 if '•' in x[1] or '•' in x[2] else 40)
    ws.cell(row=r,column=2).fill=PatternFill('solid',fgColor=BIRU)
    ws.cell(row=r,column=3).fill=PatternFill('solid',fgColor=JINGGA)
    r+=1
ws.freeze_panes='A5'
print('sheet3 ok')
wb.save(f'{OUT}/MATRIKS_OPERASIONAL_16_KELAS.xlsx')
print('SAVED tahap 1')
